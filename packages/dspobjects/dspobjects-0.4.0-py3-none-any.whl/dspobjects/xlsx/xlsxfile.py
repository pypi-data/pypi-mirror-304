"""xlsxfile.py

"""
# Package Header #
from ..header import *

# Header #
__author__ = __author__
__credits__ = __credits__
__maintainer__ = __maintainer__
__email__ = __email__


# Imports #
# Standard Libraries #
import pathlib
import warnings

# Third-Party Packages #
from baseobjects import BaseObject
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
import pandas

# import xlrd  # not in this code but needed for pandas

# Local Packages #


# Definitions #
# Functions#
def range2dataframe(worksheet: Worksheet, range_: str) -> pandas.DataFrame:
    """Converts a range of an Excel sheet into a Pandas dataframe.

    Args:
        worksheet: The worksheet which the range is located.
        range_: The range to convert into a dataframe. It must be in the Excel standard format.

    Returns:
        The Pandas dataframe of the Excel range.
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_)
    data = worksheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True)
    headers = next(data)
    return pandas.DataFrame(data, columns=headers)


# Classes #
class XLSXFile(BaseObject):
    """A class for reading Excel files.

    Attributes:
        _path: Path of Excel file.
        op_workbook: An openpyxl object of the Excel Workbook.

        op_worksheets: A dictionary of openpyxl worksheets with the names as keys
        worksheets: A dictionary of Pandas dataframes containing the Excel worksheets.

        table_ranges: A dictionary of dictionaries with table ranges of each of the Worksheets.
        tables: A dictionary of dictionaries with Pandas dataframes containing the table of each worksheet.

    Args:
        path: Path of Excel file.
        load: Determines if the Excel file will be loaded.
        init: Determines if this object with construct.
    """

    # Magic Methods #
    # Construction/Destruction
    def __init__(self, path: str | pathlib.Path | None = None, load: bool = True, init: bool = True) -> None:
        # New Attributes #
        self._path: pathlib.Path | None = None
        self.op_workbook: Workbook | None = None

        self.op_worksheets: dict[str, Worksheet] = {}
        self.worksheets: dict[str, pandas.DataFrame] = {}

        self.table_ranges: dict[str, dict[str, str]] = {}
        self.tables: dict[str, dict[str, pandas.DataFrame]] = {}

        # Parent Attributes #
        super().__init__(init=init)

        if init:
            self.construct(path=path, load=load)

    @property
    def path(self) -> pathlib.Path:
        """The path to the Excel file."""
        return self._path

    @path.setter
    def path(self, value: str | pathlib.Path) -> None:
        if isinstance(value, pathlib.Path) or value is None:
            self._path = value
        else:
            self._path = pathlib.Path(value)

    @property
    def sheet_names(self) -> list[str]:
        """A list of worksheet names from Excel workbook."""
        return self.op_workbook.sheetnames

    # Instance Methods #
    # Constructors/Destructors
    def construct(self, path: str | pathlib.Path | None = None, load: bool | None = None) -> None:
        """Constructs this object.

        Args:
            path: Path of Excel file.
            load: Determines if the Excel file will be loaded.
        """
        if path is not None:
            self.path = path

        if load and self.path is not None:
            self.load()

    def load(self, path: str | pathlib.Path | None = None) -> None:
        """Loads the data from the Excel sheet into its corresponding attributes.

        Args:
            path: Override current attribute path and the load from the new path.
        """
        self.load_workbook(path=path)
        self.load_worksheet()
        self.load_tables()

    def load_workbook(self, path: str | pathlib.Path | None = None) -> None:
        """Loads the data from the Excel sheet as an openpyxl Workbook and assigns it to op_workbook.

        Args:
            path: Override current attribute path and the load from the new path.
        """
        if path is not None:
            self.path = path
        warnings.filterwarnings("ignore")
        self.op_workbook = openpyxl.load_workbook(self.path)
        warnings.filterwarnings("default")

    def load_worksheet(self) -> None:
        """Load the worksheets as a dictionaries where names of the worksheets are the keys."""
        for i, key in enumerate(self.sheet_names):
            self.op_worksheets[key] = self.op_workbook.worksheets[i]
            self.worksheets[key] = pandas.read_excel(self.path, index_col=None, header=None)

    def load_tables(self) -> None:
        """Loads the tables from all worksheets."""
        self.table_ranges.clear()
        self.tables.clear()
        for name, worksheet in self.op_worksheets.items():
            self.table_ranges[name] = worksheet.tables.copy()
            self.tables[name] = {n: range2dataframe(worksheet, r) for n, r in worksheet.tables.items()}

    def load_worksheet_table_ranges(self, name: str) -> dict[str, str]:
        """Loads the table ranges in a worksheet.

        Args:
            name: The name of the worksheet to get the tables from.

        Returns:
            A dictionary where the keys are the names of the tables and the values are ranges of the tables.
        """
        return self.op_worksheets[name].tables.copy()

    def load_worksheet_pd_tables(self, name: str) -> dict[str, pandas.DataFrame]:
        """Loads the tables from a worksheet as a dictionary of Pandas Dataframes.

        Args:
            name: The name of the worksheet to get the tables from.

        Returns:
            A dictionary where the keys are the names of the tables and the values are Pandas Dataframes of the tables.
        """
        worksheet = self.op_worksheets[name]
        return {n: range2dataframe(worksheet, r) for n, r in worksheet.tables.items()}
