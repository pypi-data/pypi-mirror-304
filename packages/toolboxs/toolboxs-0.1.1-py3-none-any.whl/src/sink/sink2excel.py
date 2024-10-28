import io
import os
import re
import imghdr
import shutil
import random
import aiohttp
import asyncio
import requests
import colorsys
import aiofiles
from tqdm import tqdm
from enum import Enum
from io import BytesIO
from openpyxl import Workbook
from aiohttp import TCPConnector
from typing_extensions import Self
from openpyxl.comments import Comment
from tqdm.asyncio import tqdm_asyncio
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Protection
from openpyxl.styles import Border, Side,Font,PatternFill
from typing import List,Generator,final,Union,Optional,Literal,ClassVar
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from src.utils.config import Logger,Utils
from src.network.request_session import Request,FileType


class CellSkin():

    '''
        @description: Cell 样式设置
        @params: highlight_color 高亮颜色
        @params: hyperlink_color 超链接
        @author: MR.阿辉
        @datetime: 2024-02-23 07:17:13
        @return: 
    '''
    def __init__(self,
            highlight_color:str='C0504D',
            hyperlink_color:str='548dd5'
        ) -> None:
        self.highlight_color = highlight_color
        self.hyperlink_color = hyperlink_color
    
    def cell_skin(self):
        pass
    
    def header_skin(self
        
        ):
        pass

# TODO：数据类型
class CellDataType(Enum):
    TEXT = 1 # 字符串
    PIC = 2 # 图片
    NUMBER = 3 # 数值
    DATETIME = 4 # 日期

'''
    @description: 对齐方式
        水平对齐方式常见的有
            两端对齐（justify）、
            填满对齐（fill）、
            左对齐（left）、
            一般对齐（general）、
            右对齐（right）、
            居中对齐（center）、
            分散对齐（distributed）
        垂直对齐常见的有
            靠下对齐（bottom），
            居中对齐（center），
            分散对齐（distributed），
            靠上对齐（top），
            两端对齐（justify）
    @params: 
    @author: MR.阿辉
    @datetime: 2024-09-20 09:14:54
    @return: 
'''
class CellPosition(Enum):
    # 创建一个居中对齐的对象
    # horizontal='center' 水平居中
    # vertical='center' 垂直居中、vertical='center' 垂直居中、 
    # wrap_text=True 自动换行
    # indent 缩进
    
    # 水平、垂直居中，自动换行
    position_center = Alignment(horizontal='center', vertical='center',wrap_text=True)
    # 左对齐
    position_left = Alignment(horizontal='left', vertical='center',wrap_text=True,indent=1)
    # 右对齐
    position_right = Alignment(horizontal='right', vertical='center',wrap_text=True)
    # 垂直居中、水平分散对齐
    position_distributed = Alignment(horizontal='distributed', vertical='center',wrap_text=True,indent=1)



class Cell():
    
    '''
        @description: Cell 对象
        @params: val 数据值
        @params: data_type 是否为图片 True or False 默认为 False
        @params: data_format 数据格式
        @params: default 定义默认值，如果 val 为 None 就使用 default
        @params: width 设置单元格的宽度，默认为 20;可根据实际情况进行调整
        @params: position 数据在单元格中所属位置; 
        @params: image_size 设置图片宽高，默认为(90,90)
        @params: highlight 是否设置高亮，默认为False
        @params: hyperlink 超链接设置，可以不指定。
        @params: row_size 指定的行宽，默认为1，超过将做合并
        @params: col_size 指定的列宽，默认为1，超过将做合并
        @params: subrows 子列表，与当前Cell同级
        @params: subrows_direction 子列表方向，1 水平 2 垂直
        @params: font_bold 字体加粗，默认为False
        @params: font_size 字体大小 默认为11
        @params: font_name 字体名称，默认为宋体
        @params: font_color 字体颜色
        @params: fill 单元格样式
        @params: comment 单元格备注(批注)
        @params: border 边框样式
        @author: MR.阿辉
        @datetime: 2024-01-04 15:17:33
        @return: 
    '''
    def __init__(self,
            val,
            data_type:CellDataType=CellDataType.TEXT,
            data_format:str='General', # type: ignore
            default=None,
            width:float=20,
            position:CellPosition=CellPosition.position_left,
            image_size:tuple=(90,90),
            highlight:bool=False,
            hyperlink:str=None, # type: ignore
            row_size:int=1,
            col_size:int=1,
            subrows:List['Row']=None, # type: ignore
            subrows_direction:int=1,
            font_bold:bool=False,
            font_size:int=11,
            font_name:str='宋体',
            font_color:str=None, # type: ignore
            fill:PatternFill=None, # type: ignore
            comment:str=None, # type: ignore
            border:Border=None # type: ignore
        ) -> None:
        self.val = default if val is None else val
        self.data_type = data_type
        self.data_format = data_format
        self.width = width
        self.position = position
        self.image_size = image_size
        self.highlight = highlight
        self.hyperlink = hyperlink
        self.row_size = row_size
        self.col_size = col_size
        self.subrows = subrows
        self.font_bold = font_bold
        self.font_size = font_size
        self.font_name = font_name
        self.font_color = font_color
        self.fill = fill
        self.subrows_direction = subrows_direction
        self.comment = comment
        self.border = border
        
        # 高亮字体颜色
        self.highlight_color:str='C0504D'
        # 超链接字体颜色
        self.hyperlink_color:str='0563C1'
        
        # 下划线 doubleAccounting | single | singleAccounting | double
        self.underline:str = None # type: ignore


class Row():
    
    '''
        @description: 
        @params: cells 单元格集合
        @params: row_index 重定义 起始行号，默认为 -1 表示自动更新，若需要重新指定，需要设置一个大于0的数。
        @author: MR.阿辉
        @datetime: 2024-09-21 07:22:34
        @return: 
    '''
    def __init__(self, *cells:Cell,row_index:int=-1) -> None:
        self.cells = list(cells)
        self.row_index = row_index
    
    def get_cells(self):
        for cell in self.cells:
            yield cell
    
    def cell_size(self):
        return len(self.cells)
    
    def cell_insert(self,index:int,cell:Cell):
        self.cells.insert(index,cell)
        
    def get_row_index(self):
        return self.row_index

class Tools():
    
    '''
        @description: 随机颜色
        @params: 
        @author: MR.阿辉
        @datetime: 2024-02-22 20:34:37
        @return: 
    '''
    def generate_random_color(self):
        h= random.random()
        s = random.random()
        v = random.random()
        r, g,b= colorsys.hsv_to_rgb(h, s, v)
        r = int(r * 255)
        g= int(g * 255)
        b= int(b * 255)
        return f"#{r:02x}{g:02x}{b:02x}"

class Sink2Excel(Tools):

    '''
        @description: 
        @params: file_path 可以为空，但是得需要保证文件所诉目录已经存在
        @params: freeze_headers 是否冻结表头，默认为True
        @author: MR.阿辉
        @datetime: 2024-01-04 11:17:05
        @return: 
    '''
    def __init__(self,freeze_headers:bool=True): # type: ignore
        # sheet 索引
        self.sheet_index = 0
        # 冻结表头
        self.freeze_headers = freeze_headers
        # 单元格默认高度
        self.default_row_height = 60
    
    '''
        @description: 创建上下文管理器
        @params: 
        @author: MR.阿辉
        @datetime: 2024-03-22 09:01:07
        @return: 
    '''
    def __enter__(self) -> Self:
        #  实例化一个 Workbook
        self.wb = Workbook()
        # header,主要针对 请求图片时的网络设置
        self.headers= Utils.get_headers('PC-Mac')
        return self
    
    
    '''
        @description: 创建一个新的sheet
        @params: sheet_name sheet名称
        @author: MR.阿辉
        @datetime: 2024-01-05 08:58:45
        @return: 
    '''
    def __create_sheet(self,sheet_name):
        current_sheet = self.wb.create_sheet(sheet_name,self.sheet_index)
        # ws.sheet_format.defaultRowHeight = self.default_row_height

        self.sheet_index += 1
        return current_sheet
    
    '''
        @description: 表头样式设置
        @params: headers
        @author: MR.阿辉
        @datetime: 2024-02-22 20:06:01
        @return: 
    '''
    def __header_style(self,headers:List[Row]):
        for row in headers:
            for cell in row.get_cells():
                # 设置字体颜色
                cell.font_color = 'FFFFFF'
                # 设置字体格式
                cell.font_name='Microsoft YaHei'
                # 设置字体size
                cell.font_size=10
                # 设置字体背景
                cell.fill = PatternFill("solid", fgColor="262626")
                # 设置字体加粗
                cell.font_bold=True
                # 居中显示
                cell.position=CellPosition.position_center
                
                # 设置边框样式
                cell.border = Border(left=Side(border_style='thin', color='FF545454'),  # 左边框，红色
                right=Side(border_style='thin', color='FF545454'), # 右边框，红色
                top=Side(border_style='thin', color='FF545454'),   # 上边框，红色
                bottom=Side(border_style='thin', color='FF545454')) # 下边框，红色
                
                # 判断是否有子列
                if cell.subrows is not None and len(cell.subrows) > 0:
                    cell.subrows = self.__header_style(cell.subrows)
        return headers
    '''
        @description: 单元格样式设置
        @params: c openpyxl Cell 对象
        @params: cell 自定义的 Cell 对象
        @params: group_index 单元组
        @params: gradient 是否开启行渐变色
        @author: MR.阿辉
        @datetime: 2024-02-22 20:30:39
        @return: 
    '''
    def __cell_style(self,
            c,
            cell:Cell,
            group_index:int,
            gradient:bool
        ):
        # 设置边框
        border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'))
        c.border = border
        
        # 设置渐变行
        if group_index %2 == 0 and gradient:
            cell.fill = PatternFill("solid", fgColor='f2f2f2')
        
        
        # 设置高亮显示
        if cell.highlight:
            # 设置字体颜色
            cell.font_color=cell.highlight_color
            cell.fill = PatternFill("solid", fgColor="FFEBCD")
        
        # 设置超链接
        if cell.hyperlink is not None:
            c.hyperlink = cell.hyperlink
            cell.font_color= cell.hyperlink_color
            # 设置下划线
            cell.underline='single'
        
        # 设置字体样式
        c.font = Font(
            name=cell.font_name,
            size=cell.font_size,
            bold=cell.font_bold,
            color=cell.font_color,
            underline=cell.underline # type: ignore
        )
        
        # 设置边框样式
        if cell.border is not None:
            c.border = cell.border
        
        # 设置单元格样式
        if  cell.fill is not None:
            c.fill = cell.fill
        
        
        # 单元格展示形式， 居中 或 靠左靠右，默认自动换行
        c.protection = Protection(locked=True, hidden=True)
        c.alignment = cell.position.value
    
    '''
        @description: 将数据写入到 Excel中
            对于使用者来说，只需要关注 rows、headers、sheet_name 这三个参数即可，其他都是提供给递归时使用的。
        @params: rows 行信息，数据类型必须为 Cell 数组
        @params: headers 表头 数据类型必须为 Cell 数组
        @params: sheet_name sheet 名称，默认为 sheet
        @params: current_sheet 当前 sheet 
        @params: start_cell_index
        @params: start_row_index
        @params: start_group_index
        @params: gradient 是否开启 单元格的渐变
        @params: sequence 设置序列，可能会与 Row.row_index 冲突(因为可能会涉及到单元格合并)，两者不能同时开启。
        @author: MR.阿辉
        @datetime: 2024-01-05 16:35:55
        @return: 
    '''
    def write(self,
            rows:Union[List[Row],Generator[List[Row],None,None]],
            headers:Row=None, # type: ignore
            sheet_name:str='sheet',
            current_sheet=None,
            start_cell_index:int=0,
            start_row_index:int=0,
            start_group_index:int=0,
            gradient:bool=True,
            sequence:bool=False
        ):
        
        # 创建 sheet
        if current_sheet is None:
            current_sheet = self.__create_sheet(sheet_name)
        
        row_index = start_row_index
        # 生成表头，重新初始化 行号，默认从1开始
        if headers is not None and headers.cell_size() > 0:
            # 获取单元格最大深度，用于合并序列号单元格
            header_depth =max(headers.get_cells(),key = lambda x:x.row_size).row_size
            
            # 添加序号列
            if sequence:
                headers.cell_insert(0,Cell(val='序号',row_size=header_depth))
            
            row_index,_ = self.write(
                rows=self.__header_style([headers]), # 设置表头样式
                current_sheet=current_sheet,
                gradient=False # 关闭单元格渐变
            )
            # 冻结表头
            if (headers is not None  and headers.cell_size() >0) and self.freeze_headers:
                current_sheet.freeze_panes = f'A{row_index+1}'

            row_index = row_index
            
        def convert_webp_to_png(webp_bytes):
            from PIL import Image
            byte_stream = BytesIO(webp_bytes)
            # 将 WEBP 格式转换成 JPG格式
            im = Image.open(byte_stream).convert('RGB')
            
            bytes = BytesIO()
            im.save(bytes, 'JPEG')
            
            return bytes
        
        
        # TODO: 合并单元格
        def merge_cells(cell,row_index,cell_index,depth):
            # 合并单元格
            if cell.row_size > 1 or cell.col_size > 1:
                depth = max(depth,row_index+cell.row_size)-1
                
                start_row,end_row = row_index, depth
                
                next_col = cell_index+1
                start_column,end_column = next_col, next_col+cell.col_size-1
                current_sheet.merge_cells(
                    start_row = start_row,
                    end_row = end_row,
                    start_column = start_column,
                    end_column = end_column
                )
                
                
                #  设置合并单元格的宽度
                for e in range(start_column,end_column+1):
                    current_sheet.column_dimensions[get_column_letter(e)].width = cell.width+0.26
                
                
                cell_index = end_column -1
            return cell_index,depth
        
        # TODO：填充图像到单元格中
        def insert_image(cell_index:int,depth:int,max_retry_size:int=0):
            c = current_sheet.cell(row=row_index, column=cell_index+1)
            # 设置单元格样式
            self.__cell_style(c,cell,group_index,gradient)
            # 设置单元格批注
            if cell.comment is not None:
                comment = Comment(cell.comment, "Author")
                #comment.width = 300
                #comment.height = 50
                c.comment = comment
            
            
            # 合并单元格
            a,depth =merge_cells(cell,row_index,cell_index,depth)
            if max_retry_size > 1:
                return (a,depth)
            
            try:
                val:str= str(cell.val)
                if val.startswith('http'):
                    # 读取网络图片
                    response = requests.get(url=val,verify=False,headers=self.headers,timeout=15)
                    # 查看图片格式
                    image_format = imghdr.what(None, h=response.content)
                    if image_format == 'webp':
                        img = Image(convert_webp_to_png(response.content))
                    else:
                        image_bytes = BytesIO() 
                        image_bytes.write(response.content) 
                        image_bytes.seek(0)
                        img = Image(image_bytes)
                    
                elif Utils.is_file_path(val):
                    # 读取本地图片
                    img = Image(val)
                
                specs:int = 20000
                
                # 将图片写到一个单元格中，此操作会调整单元格的宽高
                if cell.col_size == 1 and cell.row_size == 1:
                    # 这两个属性分别是对应添加图片的宽高
                    img.width,img.height = cell.image_size # type: ignore
                    img.alignment = {'horizontal': 'center', 'vertical': 'middle'}   # type: ignore  水平、垂直都居中
                    # 设置 单元格高度
                    current_sheet.row_dimensions[row_index].height = img.height*0.8 # type: ignore
                    # 设置 单元格宽度
                    current_sheet.column_dimensions[get_column_letter(cell_index+1)].width = img.width * 0.15 # type: ignore
                    # 右下角固定到另一个单元格
                    # AnchorMarker(列，微调，行，微调)
                    to = AnchorMarker(cell_index+1, -specs, row_index,-specs) # 计算 列位置，创建锚标记对象,设置图片所占的row 从而确认了图片位置
                else:
                    # 右下角固定到另一个单元格
                    # AnchorMarker(列，微调，行，微调)
                    to = AnchorMarker(a+1, -specs,  depth, -specs) # 计算 列位置，创建锚标记对象,设置图片所占的row 从而确认了图片位置
                
                #合并多个单元格，并将图像保存在合并后的单元格中，
                # 左上角固定到一个单元格，
                _from = AnchorMarker(cell_index, specs, row_index-1, specs) # 计算行位置 从0开始，建锚标记对象,设置图片所占的row
                
                # 将锚标记对象设置图片对象的锚属性,图形就具备了所在位置
                img.anchor = TwoCellAnchor('twoCell', _from,to) # type: ignore
                
                # 将图片添加到 Excel 单元格中
                current_sheet.add_image(img)
                
                # ws.add_image 无法设置超链接
                
            except Exception as e:
                print(cell.val,e) 
                # 重试一次
                insert_image(cell_index,depth,max_retry_size+1)
            return (a,depth)
        
        # TODO：填充文本内容到单元格中
        def insert_value(cell_index:int,depth:int,group_index):
            
            # 生成一个 单元格
            c = current_sheet.cell(
                row=row_index, 
                column=cell_index+1
            )
            if cell.data_type == CellDataType.TEXT:
                c.value = cell.val
            elif cell.data_type == CellDataType.NUMBER:
                # 设置单元格格式
                # 判断是否为整数
                v = str(cell.val)
                if (Utils.is_number(v)):
                    c.value = float(v) if '.' in v else  int(v)
            
            # 设置单元格样式
            c.number_format=cell.data_format
            
            
            
            
            # 设置单元格批注
            if cell.comment is not None:
                comment = Comment(cell.comment, "Author")
                #comment.width = 300
                #comment.height = 50
                c.comment = comment
            
            # 设置单元格宽度
            current_sheet.column_dimensions[get_column_letter(cell_index+1)].width = cell.width+0.26
            
            # 设置单元格样式
            self.__cell_style(c,cell,group_index,gradient)
            
            # 合并单元格
            cell_index,depth =merge_cells(cell,row_index,cell_index,depth)
                
            return (cell_index,depth)
        
        
        cell_index = -1 # 防止 local variable 'cell_index' referenced before assignment
        # TODO 向 excel 中插入数据
        # 创建一个进度条，在控制台展示进度
        with tqdm(rows,desc='excel processing...',colour=super().generate_random_color()) as e:
            group_index = start_group_index
            # 行
            for serial,row in  enumerate(e):
                # 设置序号
                if sequence:
                    row.cell_insert(0,Cell(
                        val=serial+1,
                        font_bold=True,
                        position=CellPosition.position_center,
                        col_size=1,
                        data_type=CellDataType.NUMBER
                    ))
                if row.get_row_index() >0:
                    row_index = row.get_row_index()
                else:
                    row_index +=1
                # 记录当前集合中最大占用行数
                max_row_size = -1
                
                cell_index = start_cell_index
                # 列
                for cell in row.get_cells():
                    
                    # 判断是否为图片
                    if cell.data_type == CellDataType.PIC:
                        cell_index,depth = insert_image(cell_index,row_index)
                    else:
                        cell_index,depth = insert_value(cell_index,row_index,group_index=group_index+1)
                    
                    # 返回最大值
                    max_row_size = max(max_row_size,depth)
                    
                    cell_index += 1
                    # 是否包含子列
                    if cell.subrows is not None and len(cell.subrows) > 0:
                        # 判断子列的方向，1为水平，2为垂直
                        if cell.subrows_direction == 2:
                            # 垂直方向
                            sub_row_cell_index = cell_index - cell.col_size 
                            sub_row_row_index = row_index
                        elif cell.subrows_direction == 1:
                            # 水平方向
                            sub_row_cell_index = cell_index
                            sub_row_row_index = row_index-1
                            
                        sub_row_size,sub_cell_size = self.write(
                            rows=cell.subrows,
                            current_sheet=current_sheet,
                            start_cell_index=sub_row_cell_index,
                            start_row_index=sub_row_row_index, # 
                            start_group_index=group_index,
                            gradient=gradient
                        )
                        cell_index=sub_cell_size
                        
                        # 返回最大值
                        max_row_size = max(max_row_size,sub_row_size)
                
                # 一个完整的行写入成功，为一组，其实就是 row id，但是 row 一定被上面定义了。
                group_index += 1
                
                
                # 合并序号单元格
                if sequence:
                    current_sheet.merge_cells(
                        start_row = row_index ,
                        end_row = max_row_size,
                        start_column = 1,
                        end_column = 1
                    )
                row_index = max_row_size
        
        # 返回 写入的行数以及列数
        return row_index,cell_index
    
    '''
        @description: 保存到到文件中
        @params: file_path 文件地址
        @author: MR.阿辉
        @datetime: 2024-07-05 14:49:39
        @return: file_path 保存的文件地址
    '''
    def sink_file(self,file_path:str):
        self.wb.save(file_path)
        return file_path
    
    '''
        @description: 将 excel 输出到 IO 数据流中
        @params: 
        @author: MR.阿辉
        @datetime: 2024-07-05 14:51:50
        @return: 
    '''
    def sink_stream(self):
        stream = io.BytesIO()
        self.wb.save(stream)
        return stream
    
    '''
        @description: 退出上下文管理器
        @params: 
        @author: MR.阿辉
        @datetime: 2024-03-22 09:04:56
        @return: 
    '''
    def __exit__(self,exc_type,exc_value,exc_traceback):
        # 释放资源
        self.wb.close()

class Unit():
    '''
        @description: 
        @params: line_no 所在 index.m3u8 的行号
        @params: invalid 是否是无效数据，例如 以#开头的
        @author: MR.阿辉
        @datetime: 2024-03-02 17:18:44
        @return: 
    '''
    def __init__(self,
                line:str,
                source_uri:str='',
                full_url:str='',
                is_key:bool=False,
                line_no:int=0,
                invalid:bool=False
            ) -> None:
        self.full_url = full_url
        self.source_uri = source_uri
        self.line = line
        self.is_key = is_key
        self.save_path:str = ''
        self.line_no = line_no
        self.invalid = invalid
    
    def __str__(self):
        return f'Unit(line={self.line},source_uri={self.source_uri},full_url={self.full_url},is_key={self.is_key},line_no={self.line_no},invalid={self.invalid})'



class Sink2Video(Logger):
    # TODO: ffmpeg 安装目录
    FFMPEG_HOME:ClassVar[str] = '/Users/zhangzhonghui/Documents/plug/ffmpeg'
    
    '''
        @description: 
        @params: m3u8_path m3u8 地址，可以是 url地址也可以是 文件路径
        @params: max_retry_size 任务下载异常重最大试次数
        @params: semaphore_size 并发控制
        @params: waiting 等待时长，下载异常，等待多少秒之后重新下载
        @author: MR.阿辉
        @datetime: 2024-03-01 17:09:08
        @return: 
    '''
    def __init__(self,
                m3u8_path:str,
                max_retry_size:int=3,
                semaphore_size:int=10,
                waiting:int=3,
                project_file:Optional[str]=None
            ) -> None:
        project_name,_,source_dir = Utils.get_project_info(project_file)
        
        super().__init__(log_file_path=f'{source_dir}/log/video',log_file_name=project_name,logger_name='video') 
        self.m3u8_path = m3u8_path
        self.max_retry_size = max_retry_size
        self.semaphore_size = semaphore_size
        self.waiting = waiting
        # m3u8是否加密
        self.is_encryption = False
        self.request = Request(project_file)
    
    '''
        @description: 解析 index.m3u8 文件
        @params: callback 回调函数，解析规则由用户定义
        @author: MR.阿辉
        @datetime: 2024-03-02 09:19:15
        @return: 
    '''
    def __parse_ts(self,callback):
        # 读取 m3u8 文件，一行一行的向用户返回
        def parse_m3u8():
            if Utils.is_url(self.m3u8_path):
                def callback(resp):
                    yield from resp.text.split('\n')
                    
                yield from self.request.send(url=self.m3u8_path)(callback)
            else:
                with open(self.m3u8_path,mode='r') as f:
                    yield from f.readlines()
        
        for index,line in  enumerate(parse_m3u8()):
            # 去掉字符串两边的空格以及最后的换行符
            line = line.rstrip("\n")
            
            # ts 数据行
            unit = Unit(line = line,source_uri=line,is_key=False,line_no=index)
            
            # 获取 key 的 uri 地址
            if len(re.findall('EXT-X-KEY',line)) > 0:
                self.is_encryption = True
                key_uri = re.findall('URI="(.*?)"',line,re.IGNORECASE)[0]
                unit = Unit(line = line,source_uri = key_uri,is_key=True,line_no=index)
            elif line.startswith('#'):
                #无效数据
                unit = Unit(line = line,is_key=False,line_no=index,invalid=True)
            yield callback(unit)
    
    '''
        @description: 主程序，用于构建下载 ts文件的协任务
        @params: ts_save_path ts 文件保存地址
        @params: callback 处理 m3u8.index 文件的回调函数
        @params: mode 视频写入模式，a 追加写，支持断点续传，w 覆盖写，从头开始下载。默认为 a
        @author: MR.阿辉
        @datetime: 2024-03-01 17:15:53
        @return: 返回当前对象
    '''
    def run(self,
            ts_save_path:str,
            callback,
            mode: Literal['a','w']='a'
        ) -> Self:
        
        '''
            @description: 检查 ts 文件是否存在
            @params: file_path 文件地址
            @author: MR.阿辉
            @datetime: 2024-03-13 08:52:43
            @return: bool 类型 存在返回True，否则为False
        '''
        async def ts_exists(file_path:str) -> bool:
            return os.path.isfile(file_path)
        
        '''
            @description: 数据下载
            @params: semaphore 并发控制
            @params: url 请求地址
            @params: save_path 保存地址
            @author: MR.阿辉
            @datetime: 2024-03-01 17:21:40
            @return: 
        '''
        async def download(semaphore,unit:Unit,retry_size:int=0):
            # 如果为无效数据，或超过最大重试次数直接返回
            if unit.invalid or retry_size>= self.max_retry_size:
                return unit
            url,save_path = unit.full_url,unit.save_path
            
            
            try:
                # 并发控制
                async with semaphore:
                    # 是否选择追加写入模式，并且检查文件是否存在
                    if(mode == 'a' and await ts_exists(save_path)):
                        self.logger.info(f'{save_path} Already exists, skipping...')
                        return unit
                    
                    self.logger.info('准备请求: '+unit.full_url)
                    # 超时设置
                    timeout = aiohttp.ClientTimeout(total=60)
                    async with aiohttp.ClientSession(headers=self.request.session.headers,connector=TCPConnector(ssl=False),timeout=timeout) as session: # type: ignore
                        async with session.get(url=url) as resp:
                            if int(resp.status) >= 200 and int(resp.status) <300:
                                async with aiofiles.open(save_path,mode='wb') as f:
                                    await f.write(await resp.read())
                                    return unit
                            else:
                                self.logger.info('异常状态:',resp.status,f'暂停 {self.waiting} 秒后重试...')
                                
                                
            except asyncio.TimeoutError as e:
                self.logger.exception(f"请求超时:\n{e}")
            except aiohttp.ClientResponseError as e:
                self.logger.exception(f"请求异常:\n{e}")
            except Exception as e:
                self.logger.exception(f"未知异常:\n{e}")
            
            # 下一次重新次数
            next_retry_size = retry_size+1
            self.logger.info(f'当前执行节点：\n{unit} \n等待{self.waiting} 秒后，尝试第{next_retry_size}次重试...')
            await asyncio.sleep(self.waiting)
            return await download(semaphore,unit,retry_size=next_retry_size)
        '''
            @description:  构建下载 ts 文件 的协程任务
            @params: save_path ts 保存文件目录
            @params: units Unit 对象集,封装了单个ts数据信息
            @author: MR.阿辉
            @datetime: 2024-03-01 17:18:34
            @return: 
        '''
        async def build_task(save_path:str,units:Generator[Unit,None,None]):
            
            # 并发控制
            semaphore = asyncio.Semaphore(self.semaphore_size)
            
            tasks = []
            index=-1
            for unit in units:
                if not unit.invalid:
                    index += 1
                if  unit.is_key:
                    # 构建保存key的文件路径
                    unit.save_path = self.request.save_path(file_path=save_path,file_name=f'key',file_type=FileType.KEY)
                elif not unit.invalid:
                    # 构建保存ts的文件路径
                    unit.save_path = self.request.save_path(file_path=save_path,file_name=f'{index}',file_type=FileType.TS)
                # 创建协程task
                tasks.append(asyncio.create_task(download(semaphore,unit)))
            
            return await tqdm_asyncio.gather(*tasks,desc='downloading',unit='bar')
        
        '''
            @description: task 回调函数，对数据进行二次加工
            @params: unit
            @author: MR.阿辉
            @datetime: 2024-03-04 10:21:30
            @return: 
        '''
        def task_callback(unit:Unit) -> Unit:
            if not unit.invalid:
                if unit.is_key:
                    if self.is_encryption:
                        unit.line = unit.line.replace(unit.source_uri,'key.key')
                    else:
                        unit.line = unit.line.replace(unit.source_uri,unit.save_path)
                else:
                    if self.is_encryption:
                        unit.line = os.path.basename(unit.save_path)
                    else:
                        unit.line = f"file '{unit.save_path}'"
            return unit
        
        
        unit_list = asyncio.run(build_task(ts_save_path, self.__parse_ts(callback)))
        
        self.logger.info('生成 index.m3u8文件并保存到文件中...')
        # 重新保存进行文件夹中
        self.finl_m3u8_file:str = self.request.save_path(file_path=ts_save_path,file_name='index.m3u8',file_type=FileType.M3U8)
        self.logger.info(f'index.m3u8 文件地址: {self.finl_m3u8_file}')
        with open(self.finl_m3u8_file,mode='w') as f:
            sorted_unit_list = sorted(list(map(task_callback,unit_list)),key= lambda x:x.line_no)
            lines = list(map(lambda x:x.line,sorted_unit_list))
            f.write('\n'.join(lines))

        return self
    
    '''
        @description: 视频合并
        @params: file_name 视频名称
        @params: out_path 视频输出地址
        @params: clearn 是否清除历史数据
        @author: MR.阿辉
        @datetime: 2024-03-02 20:26:59
        @return: 
    '''
    def merge(self,file_name:str,out_path:str,clearn:bool=True) -> Self:
        
        try:
            outfile_path:str = self.request.save_path(file_path=out_path,file_name=file_name,file_type=FileType.MP4)
            if self.is_encryption:
                md = f'{Sink2Video.FFMPEG_HOME} -allowed_extensions ALL -protocol_whitelist "file,http,crypto,tcp" -i {self.finl_m3u8_file} -c copy {outfile_path}'
            else:
                md = f"{Sink2Video.FFMPEG_HOME} -f concat -safe 0 -i {self.finl_m3u8_file} -c copy {outfile_path}"
            
            self.logger.info('ffmpeg 运行命令:',md)
            os.system(md)
        finally:
            # 清空ts文件夹下所有内容
            if clearn:
                ts_dir = os.path.dirname(self.finl_m3u8_file)
                self.logger.info('历史资源清理:',ts_dir)
                shutil.rmtree(ts_dir)
        return self