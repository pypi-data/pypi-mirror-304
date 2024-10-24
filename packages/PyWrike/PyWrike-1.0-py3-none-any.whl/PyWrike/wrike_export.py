import requests
import json
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import pandas as pd
import os
from PyWrike.gateways import OAuth2Gateway1

# Function to validate the access token
def validate_token(wrike_api_token):
    endpoint = 'https://www.wrike.com/api/v4/contacts'
    headers = {
        'Authorization': f'Bearer {wrike_api_token}'
    }
    
    response = requests.get(endpoint, headers=headers)
    if response.status_code == 200:
        print("Access token is valid.")
        return True
    else:
        print(f"Access token is invalid. Status code: {response.status_code}")
        return False

# Function to authenticate using OAuth2 if the token is invalid
def authenticate_with_oauth2(client_id, client_secret, redirect_url):
    wrike = OAuth2Gateway1(client_id=client_id, client_secret=client_secret)
    
    # Start the OAuth2 authentication process
    auth_info = {
        'redirect_uri': redirect_url
    }
    
    # Perform OAuth2 authentication and retrieve the access token
    wrike_api_token = wrike.authenticate(auth_info=auth_info)
    
    print(f"New access token obtained: {wrike_api_token}")
    return wrike_api_token

# Cache for user details
user_cache = {}

# Helper function to create headers
def create_headers(wrike_api_token):
    return {
        "Authorization": f"Bearer {wrike_api_token}"
    }

# Retry mechanism for handling rate limits
def retry_request(url, headers, retries=3, delay=60):
    for _ in range(retries):
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response
        elif response.status_code == 429:
            print("Rate limit exceeded. Sleeping for 60 seconds...")
            time.sleep(delay)
        else:
            response.raise_for_status()
    raise Exception(f"Failed after {retries} retries")

# Function to get all spaces
def get_all_spaces(wrike_api_token):
    url = 'https://www.wrike.com/api/v4/spaces'
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print("Fetching all spaces...")
    
    try:
        return response.json()["data"]
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Function to get the space ID from space name
def get_space_id_from_name(space_name, spaces):
    for space in spaces:
        if space["title"] == space_name:
            return space["id"]
    return None

# Function to get all folders and subfolders in the space
def get_all_folders(space_id, wrike_api_token):
    url = f'https://www.wrike.com/api/v4/spaces/{space_id}/folders'
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print("Fetching all folders and subfolders...")
    
    try:
        return response.json()
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Function to get folder by ID
def get_folder_by_id(folder_id, folders):
    for folder in folders:
        if folder["id"] == folder_id:
            return folder
    return None

# Function to get titles hierarchy with IDs
def get_titles_hierarchy(folder_id, folders, path=""):
    folder = get_folder_by_id(folder_id, folders)
    if not folder:
        return []
    
    current_path = f"{path}/{folder['title']}"
    current_entry = {"id": folder_id, "path": current_path}
    paths = [current_entry]
    
    for child_id in folder.get("childIds", []):
        child_paths = get_titles_hierarchy(child_id, folders, current_path)
        paths.extend(child_paths)
    
    return paths

# Function to get task details by ID with custom status mapping
def get_task_details(task_id, wrike_api_token, custom_status_mapping, custom_field_mapping):
    url = f'https://www.wrike.com/api/v4/tasks/{task_id}?fields=["effortAllocation"]'
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print(f"Fetching details for task {task_id}")
    
    try:
        task_data = response.json()["data"][0]
        custom_status_id = task_data.get("customStatusId")
        task_data["customStatus"] = custom_status_mapping.get(custom_status_id, "Unknown")
        # Process custom fields by mapping ID to name
        custom_fields = task_data.get("customFields", [])
        custom_field_data = {custom_field_mapping.get(cf["id"], "Unknown Field"): cf.get("value", "") for cf in custom_fields}
        task_data["customFields"] = custom_field_data

        return task_data
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Function to get tasks for a folder
def get_tasks_for_folder(folder_id, wrike_api_token):
    url = f'https://www.wrike.com/api/v4/folders/{folder_id}/tasks?fields=["subTaskIds","effortAllocation","authorIds","customItemTypeId","responsibleIds","description","hasAttachments","dependencyIds","superParentIds","superTaskIds","metadata","customFields","parentIds","sharedIds","recurrent","briefDescription","attachmentCount"]'
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print(f"Fetching tasks for folder {folder_id}")
    
    try:
        return response.json()["data"]
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Recursive function to get all subtask IDs
def get_all_subtask_ids(task, token):
    task_ids = [{"id": task["id"], "title": task["title"]}]
    if "subTaskIds" in task:
        for subtask_id in task["subTaskIds"]:
            subtask = get_task_details(subtask_id, token, {})
            task_ids.extend(get_all_subtask_ids(subtask, token))
    return task_ids

# Function to clean HTML content and preserve line breaks
def clean_html(raw_html):
    soup = BeautifulSoup(raw_html, "html.parser")
    lines = soup.stripped_strings
    return "\n".join(lines)

# Function to get user details by ID
def get_user_details(user_id, wrike_api_token):
    if user_id in user_cache:
        return user_cache[user_id]
    
    url = f"https://www.wrike.com/api/v4/users/{user_id}"
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print(f"Fetching details for user {user_id}")
    
    try:
        user_data = response.json()["data"][0]
        email = user_data["profiles"][0]["email"]
        user_cache[user_id] = email
        return email
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Function to get custom statuses
def get_custom_statuses(wrike_api_token):
    url = 'https://www.wrike.com/api/v4/workflows'
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print("Fetching custom statuses...")
    
    try:
        return response.json()["data"]
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Function to create a mapping from customStatusId to custom status name
def create_custom_status_mapping(workflows):
    custom_status_mapping = {}
    for workflow in workflows:
        for status in workflow.get("customStatuses", []):
            custom_status_mapping[status["id"]] = status["name"]
    return custom_status_mapping

# Function to get custom statuses
def get_custom_fields(wrike_api_token):
    url = 'https://www.wrike.com/api/v4/customfields'
    headers = create_headers(wrike_api_token)
    response = retry_request(url, headers=headers)
    print("Fetching custom fields...")
    
    try:
        return response.json()["data"]
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        print(f"Response content: {response.content}")
        raise

# Create a mapping from customFieldId to customFieldName and customFieldType
def create_custom_field_mapping(custom_fields):
    custom_field_mapping = {}
    for field in custom_fields:
        field_title = field["title"]
        field_type = field["type"]
        # Store both name and type in the mapping
        custom_field_mapping[field["id"]] = f"{field_title} [{field_type}]"
    return custom_field_mapping

def wrike_export_main():
    # Load configuration and delete task details from Excel file
    excel_file = input("Enter the path to the Excel file: ")
    if not os.path.isfile(excel_file):
        print("File does not exist. Please check the path.")
        exit()

    try:
        config_df = pd.read_excel(excel_file, sheet_name="Config", header=1)
    except Exception as e:
        print(f"Failed to read Excel file: {e}")
        exit()

    # Extract information from the Config sheet 
    wrike_api_token = config_df.at[0, "Token"]
    space_name = config_df.at[0, "Space to extract data from"]

        # Validate the token
    if not validate_token(wrike_api_token):
        # If the token is invalid, authenticate using OAuth2 and update the access_token
        wrike = OAuth2Gateway1(excel_filepath=excel_file)
        wrike_api_token = wrike._create_auth_info()  # Perform OAuth2 authentication
        print(f"New access token obtained: {wrike_api_token}")

    # Fetch all spaces and get the space ID for the given space name
    spaces_response = get_all_spaces(wrike_api_token)
    space_id = get_space_id_from_name(space_name, spaces_response)

    if not space_id:
        print(f"Space with name '{space_name}' not found.")
        exit()

    # Extract titles hierarchy for all top-level folders
    print("Extracting titles hierarchy for all top-level folders...")
    folders_response = get_all_folders(space_id, wrike_api_token)
    all_paths = []
    processed_folder_ids = set()
    for folder in folders_response["data"]:
        if "scope" in folder and folder["scope"] == "WsFolder":
            paths = get_titles_hierarchy(folder["id"], folders_response["data"])
            all_paths.extend(paths)

    # Fetch custom statuses and create a mapping
    workflows_response = get_custom_statuses(wrike_api_token)
    custom_status_mapping = create_custom_status_mapping(workflows_response)

    # Fetch custom fields and create a mapping
    custom_fields_response = get_custom_fields(wrike_api_token)
    custom_field_mapping = create_custom_field_mapping(custom_fields_response)

    # Initialize a set to store all unique custom field names
    custom_field_names = set()

    # Extract tasks and subtasks for each folder
    print("Extracting tasks and subtasks for each folder...")
    all_folders_tasks = []
    task_counter = 1  # Counter for task keys
    for folder in all_paths:
        folder_id = folder["id"]
        if folder_id in processed_folder_ids:
            print(f"Skipping already processed folder {folder_id}.")
            continue
        processed_folder_ids.add(folder_id)
        print(f"Processing folder {folder_id} - {folder['path']}...")
        try:
            tasks = get_tasks_for_folder(folder_id, wrike_api_token)
            folder_tasks = {
            "folder_id": folder_id,
            "folder_path": folder["path"].replace(f"/{space_name}", "", 1) if folder["path"].startswith(f"/{space_name}/") else folder["path"].replace(f"{space_name}", ""),
            "tasks": []
    }

            print(f"Found {len(tasks)} tasks in folder {folder_id}.")
            for task in tasks:
                task_details = get_task_details(task["id"], wrike_api_token, custom_status_mapping, custom_field_mapping)
                dates = task_details.get("dates", {})
                start_date = dates.get("start", "")
                due_date = dates.get("due", "")
                duration = dates.get("duration", "")
                efforts = task_details.get("effortAllocation", {})
                effort = efforts.get("totalEffort", "")
                # Clean the HTML content for description
                description_html = task.get("description", "")
                description_cleaned = clean_html(description_html)

                # Fetch emails for responsible IDs
                responsible_emails = []
                for user_id in task_details.get("responsibleIds", []):
                    try:
                        responsible_emails.append(get_user_details(user_id, wrike_api_token))
                    except requests.exceptions.HTTPError as e:
                        print(f"Error fetching user details for {user_id}: {e}")
                        if e.response.status_code == 429:
                            print("Rate limit exceeded. Sleeping for 60 seconds...")
                            time.sleep(60)
                        responsible_emails.append("Unknown")
                    except Exception as e:
                        print(f"Unexpected error: {e}")
                        responsible_emails.append("Unknown")
                responsible_emails_str = ", ".join(responsible_emails)

                
                # Update the set of custom field names with the current task's custom fields
                for field_id, value in task_details.get("customFields", {}).items():
                    custom_field_names.add(custom_field_mapping.get(field_id, field_id))

                task_key = f"T{task_counter}"
                task_counter += 1
                folder_tasks["tasks"].append({
                    "task_key": task_key,
                    "task_id": task["id"],
                    "task_title": task["title"],
                    "task_description": description_cleaned,
                    "task_responsibleEmails": responsible_emails_str,
                    "status": task_details.get("status", ""),
                    "priority": task_details.get("importance", ""),
                    "custom_status": task_details.get("customStatus", ""),
                    "custom_fields": task_details.get("customFields", {}), 
                    "start_date": start_date,
                    "due_date": due_date,
                    "duration": duration,
                    "effort": effort,
                    "time_spent": task_details.get("timeSpent", ""),
                    "subtasks": []
                })
                subtask_counter = 1
                if "subTaskIds" in task_details:
                    for subtask_id in task_details["subTaskIds"]:
                        subtask_details = get_task_details(subtask_id, wrike_api_token, custom_status_mapping, custom_field_mapping)
                        subtask_dates = subtask_details.get("dates", {})
                        subtask_start_date = subtask_dates.get("start", "")
                        subtask_due_date = subtask_dates.get("due", "")
                        subtask_duration = subtask_dates.get("duration", "")
                        subtask_efforts = subtask_details.get("effortAllocation", {})
                        subtask_effort = subtask_efforts.get("totalEffort", "")
                        subtask_html = subtask_details.get("description", "")
                        subtask_description_cleaned = clean_html(subtask_html)

                        # Fetch emails for responsible IDs
                        subtask_responsible_emails = []
                        for user_id in subtask_details.get("responsibleIds", []):
                            try:
                                subtask_responsible_emails.append(get_user_details(user_id, wrike_api_token))
                            except requests.exceptions.HTTPError as e:
                                print(f"Error fetching user details for {user_id}: {e}")
                                if e.response.status_code == 429:
                                    print("Rate limit exceeded. Sleeping for 60 seconds...")
                                    time.sleep(60)
                                subtask_responsible_emails.append("Unknown")
                            except Exception as e:
                                print(f"Unexpected error: {e}")
                                subtask_responsible_emails.append("Unknown")
                        subtask_responsible_emails_str = ", ".join(subtask_responsible_emails)

                        for field_id, value in subtask_details.get("customFields", {}).items():
                            custom_field_names.add(custom_field_mapping.get(field_id, field_id))
                        
                        
                        folder_tasks["tasks"][-1]["subtasks"].append({
                            "subtask_key": f"{task_key}.{subtask_counter}",
                            "subtask_id": subtask_details["id"],
                            "subtask_title": subtask_details["title"],
                            "subtask_description": subtask_description_cleaned,
                            "subtask_responsibleEmails": subtask_responsible_emails_str,
                            "status": subtask_details.get("status", ""),
                            "priority": subtask_details.get("importance", ""),
                            "custom_status": subtask_details.get("customStatus", ""),
                            "subtask_custom_fields": subtask_details.get("customFields", {}),
                            "start_date": subtask_start_date,
                            "due_date": subtask_due_date,
                            "duration": subtask_duration,
                            "effort": subtask_effort,
                            "time_spent": subtask_details.get("timeSpent", "")
                        })
                        subtask_counter += 1
                print(f"Task {task['id']} ({task_details['title']}) with {len(folder_tasks['tasks'][-1]['subtasks'])} subtasks processed.")
            all_folders_tasks.append(folder_tasks)
        except requests.exceptions.HTTPError as e:
            print(f"Error fetching tasks for folder {folder_id}: {e}")
            if e.response.status_code == 429:
                print("Rate limit exceeded. Sleeping for 60 seconds...")
                time.sleep(60)
        except Exception as e:
            print(f"Unexpected error: {e}")

    # Initialize the workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks and Subtasks"

    # Define base headers
    base_headers = ["Key", "Folder", "Parent Task", "Task Title", "Status", "Priority", "Assigned To", "Custom Status", "Start Date", "Duration", "Effort", "Time Spent", "End Date", "Description"]

    # Collect all unique custom field names
    custom_field_names = set()

    # Iterate over folders and tasks
    for folder in folders_response["data"]:
        if "scope" in folder and folder["scope"] == "WsFolder":
            # Get tasks for the current folder
            tasks = get_tasks_for_folder(folder["id"], wrike_api_token)
            
            for task in tasks:
                # Get main task details and update custom field names
                task_details = get_task_details(task["id"], wrike_api_token, custom_status_mapping, custom_field_mapping)
                custom_field_names.update(task_details["customFields"].keys())
                
                # Check if the task has subtasks
                if "subTaskIds" in task:
                    for subtask_id in task["subTaskIds"]:
                        # Get subtask details and update custom field names
                        subtask_details = get_task_details(subtask_id, wrike_api_token, custom_status_mapping, custom_field_mapping)
                        custom_field_names.update(subtask_details["customFields"].keys())

    # Combine base headers with dynamic custom field headers (name and type)
    headers = base_headers + [custom_field_mapping.get(field_id, field_id) for field_id in custom_field_names]
    ws.append(headers)

    # Populate workbook with tasks and subtasks
    for folder_tasks in all_folders_tasks:
        folder_path = folder_tasks["folder_path"]
        for task in folder_tasks["tasks"]:
            
            task_data = [
                task["task_key"],
                folder_path,
                "",
                task["task_title"],
                task["status"],
                task["priority"],
                task["task_responsibleEmails"],
                task["custom_status"],
            
                task["start_date"],
                task["duration"],
                task["effort"],
                task["time_spent"],
                task["due_date"],
                task["task_description"]
            ]
            
            for field in custom_field_names:
                task_data.append(task["custom_fields"].get(field, ""))  # Use task's custom_fields here
        
            ws.append(task_data)

            for subtask in task["subtasks"]:
            
                subtask_data = [
                    subtask["subtask_key"],
                    folder_path,
                    task["task_title"],
                    subtask["subtask_title"],
                    subtask["status"],
                    subtask["priority"],
                    subtask["subtask_responsibleEmails"],
                    subtask["custom_status"],
                    
                    subtask["start_date"],
                    subtask["duration"],
                    subtask["effort"],
                    subtask["time_spent"],
                    subtask["due_date"],
                    subtask["subtask_description"]
                ]
                
                # Add custom field values for subtasks
                for field in custom_field_names:
                    subtask_data.append(subtask["subtask_custom_fields"].get(field, ""))  # Use subtask's custom_fields here

                ws.append(subtask_data)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook to a file
    output_filename = f"export_{space_name.replace(' ', '_')}.xlsx"
    wb.save(output_filename)
    print(f"Workbook '{output_filename}'")

if __name__ == "__wrike_export_main__":
    wrike_export_main()