# coding=utf-8

"""
@fileName       :   excel.py
@data           :   2024/2/22
@author         :   jiangmenggui@hosonsoft.com
"""
import os.path

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:

    def __init__(self, excel_filename: str):
        if not excel_filename.endswith('.xlsx'):
            raise TypeError('Only support xlsx file')
        if not os.path.exists(excel_filename):
            raise FileNotFoundError(f"{excel_filename} not found!")
        self._file = excel_filename
        self._excel = self._get_excel()

    def _get_excel(self):
        return load_workbook(self._file, read_only=True, data_only=True)

    def _get_sheet(self, sheet: int | None | str) -> Worksheet:
        if isinstance(sheet, int):
            return self._excel.worksheets[sheet]
        elif isinstance(sheet, str):
            return self._excel[sheet]
        return self._excel.active

    def read(self, sheet: int | None | str = None, /, *, has_title: bool = True, min_row=None, max_row=None,
             min_col=None, max_col=None):
        """
        读取Excel的Sheet页中的数据

        如果没有指定索引，则范围从A1开始。如果工作表中没有单元格，则返回一个空元组。

        :param sheet: sheet页的顺序或名字或默认
        :type sheet: int | None | str

        :param has_title: 是否包含标题，如果包含标题则每行数据都会以字典的形式返回，否则以元组的形式返回
        :type has_title: bool

        :param min_col: 最小列索引(从1开始)
        :type min_col: int

        :param min_row: 最小行索引(从1开始)
        :type min_row: int

        :param max_col: 最大列索引(从1开始)
        :type max_col: int

        :param max_row: 最大行索引(从1开始)
        :type max_row: int

        :rtype: generator
        """
        sheet = self._get_sheet(sheet)
        iter_rows = sheet.iter_rows(min_row, max_row, min_col, max_col, values_only=True)
        if has_title:
            title = next(iter_rows)
            yield from (dict(zip(title, r)) for r in iter_rows)
        else:
            yield from iter_rows

    @property
    def sheet_names(self):
        return self._excel.sheetnames

    def close(self):
        self._excel.close()


class ExcelWrite:

    def __init__(self, excel_filename: str):
        if not excel_filename.endswith('.xlsx'):
            raise TypeError('Only support xlsx file')
        self._file = excel_filename
        self._sheets: dict[str, Worksheet] = {}
        self._excel = self._get_excel()

    def _get_excel(self) -> Workbook:  # noqa
        return Workbook(write_only=True)

    def write(self, data: list[list | dict], /, *, sheet_name=None) -> Worksheet:
        if not sheet_name:
            sheet_name = f'Sheet{len(self._sheets) + 1}'
        if sheet_name not in self._sheets:
            self._sheets[sheet_name] = self._excel.create_sheet(title=sheet_name)
        if not data:
            return self._sheets[sheet_name]

        data = iter(data)
        first = next(data)
        if isinstance(first, dict):
            self._sheets[sheet_name].append(tuple(first.keys()))  # write_title
            self._sheets[sheet_name].append(tuple(first.values()))
        else:
            self._sheets[sheet_name].append(first)
        for r in data:
            if isinstance(r, dict):
                self._sheets[sheet_name].append(tuple(r.values()))
            else:
                self._sheets[sheet_name].append(r)
        return self._sheets[sheet_name]

    def save(self):
        self._excel.save(self._file)

    def close(self):
        self._excel.close()


class Excel(ExcelReader, ExcelWrite):

    def __init__(self, excel_filename: str):
        super().__init__(excel_filename=excel_filename)
        self._sheets = {name: self._excel[name] for name in self._excel.sheetnames}

    def _get_excel(self):  # noqa
        return load_workbook(self._file, data_only=True)

    def write(self, data, /, *, sheet_name=None):
        if not sheet_name:
            sheet_name = self._excel.active.title
        return super().write(data, sheet_name=sheet_name)

    def __getitem__(self, item) -> Worksheet:
        return self._excel[item]


if __name__ == '__main__':
    # excel = Excel(r"./test.xlsx")
    # sheet = excel['Sheet2']
    # print(sheet)
    # for row in excel.read(has_title=True):
    #     print(row)
    # print()
    # excel.write([{"name": 'Jack', 'age': 77}])
    # excel.save()
    pass
