import datetime
import pandas as pd
import openpyxl as xl
from ..library import nth
from ..models.database import NodeInRecord, Record
from ..models.database.library import TableControl
from ..models.line_by_line_leaf import LineByLineLeafModel
from .style import StyleControl


class TreeDrawer():
    """エクセルで罫線などを駆使して、樹形図を描画します"""


    def __init__(self, table, ws, settings_obj, debug_write=False):
        """初期化
        
        Parameters
        ----------
        table : Table
            ツリーテーブル
        ws : openpyxl.Worksheet
            ワークシート
        settings_obj : Settings
            各種設定
        debug_write : bool
            デバッグライト
            DEBUG_TIPS: デバッグライトをオンにして、コンソールにログを表示すると不具合を調査しやすくなります
        """
        self._table = table
        self._ws = ws
        self._settings_obj = settings_obj
        self._debug_write = debug_write

        self._prev_record = Record.new_empty(specified_end_th_of_node=self._table.analyzer.end_th_of_node)
        self._curr_record = Record.new_empty(specified_end_th_of_node=self._table.analyzer.end_th_of_node)
        self._next_record = Record.new_empty(specified_end_th_of_node=self._table.analyzer.end_th_of_node)


    def render(self):
        """描画"""

        # 列幅の自動調整
        # --------------
        #
        #   NOTE 'no' は列ではなくインデックスなのでやり方が異なる
        #   NOTE 文字数は取れるが、１文字の横幅が１とは限らない
        #
        # 'no' インデックス
        width = len('no')
        for index_value in self._table.df.index:
            width = max(width, len(str(index_value)))
        self._ws.column_dimensions['A'].width = width * 1.5

        # エッジ列が省略されているケース
        #
        #   NOTE 規定値を何か設定しないと、デフォルトの長い列幅になってしまう
        #
        for source_column_th, column_name in enumerate(self._table.df.columns, 1):

            # ノードを見つければ、その前列がエッジ
            result = TableControl.pattern_of_column_name_of_node.match(column_name)
            if result:
                target_column_th = StyleControl.get_target_column_th(source_table=self._table, column_name=column_name)
                target_column_letter = xl.utils.get_column_letter(target_column_th - 1)
                # 幅は 4 でいいだろ
                self._ws.column_dimensions[target_column_letter].width = 4
                continue

        # 各列
        for source_column_th, column_name in enumerate(self._table.df.columns, 1):
            target_column_th = StyleControl.get_target_column_th(source_table=self._table, column_name=column_name)
            target_column_letter = xl.utils.get_column_letter(target_column_th)
            number_of_character = StyleControl.get_number_of_character_of_column(df=self._table.df, column_name=column_name)
            # Root とか 1st とか、ヘッダーに入っているので、とりあえず最低でも 4文字は欲しい。 1000th とかはとりあえず考えないことにする
            number_of_character = max(number_of_character, 4)

            #print(f"列幅の自動調整  {column_name=}  {target_column_letter=}  内訳：  {source_column_th=}  {target_column_th=}  {number_of_character=}")

            # ノード
            result = TableControl.pattern_of_column_name_of_node.match(column_name)
            if result:
                # 文字幅を 1.2 倍 + 1 ぐらいしておく
                # FIXME フォント情報からきっちり横幅を取れないか？
                self._ws.column_dimensions[target_column_letter].width = number_of_character * 1.2 + 1
                continue


            # エッジ
            #
            #   余白を開けたいから広くとる
            #
            result = TableControl.pattern_of_column_name_of_edge.match(column_name)
            if result:
                # 文字幅の 1.2 倍 + 4 ぐらいしておく
                # FIXME フォント情報からきっちり横幅を取れないか？
                self._ws.column_dimensions[target_column_letter].width = number_of_character * 1.2 + 4
                continue

            # 余り列
            # ------
            # 余り情報だし、余白は要らないから、文字幅を 1.2 倍ぐらいしておく
            # FIXME フォント情報からきっちり横幅を取れないか？
            self._ws.column_dimensions[target_column_letter].width = number_of_character * 1.2


        # 対象シートへ列ヘッダー書出し
        self._on_header()

        # 対象シートへの各行書出し
        self._table.for_each(on_each=self._on_each_record)

        # 最終行の実行
        self._on_each_record(next_row_number=len(self._table.df), next_record=Record.new_empty(specified_end_th_of_node=self._table.analyzer.end_th_of_node))

        # ウィンドウ枠の固定
        self._ws.freeze_panes = 'B2'


    def _forward_cursor(self, next_record):
        """送り出し

        Parameters
        ----------
        next_record : Record
            次行
        """
        self._prev_record = self._curr_record
        self._curr_record = self._next_record
        self._next_record = next_record


    def _on_header(self):

        # 変数名の短縮
        ws = self._ws


        # 列の幅設定
        column_width_dict = {}

        width = self._settings_obj.dictionary['column_width_of_no']
        if width is not None:
            column_width_dict['A'] = width      # no

        width = self._settings_obj.dictionary['column_width_of_root_side_padding']
        if width is not None:
            column_width_dict['B'] = width      # B列の幅。ツリー構造図の根側パディング
        
        width = self._settings_obj.dictionary['column_width_of_node']
        if width is not None:
            column_width_dict['C'] = width      # 根


        head_column_th = 4
        for node_th in range(1, self._table.analyzer.end_th_of_node):

            width = self._settings_obj.dictionary['column_width_of_parent_side_edge']
            if width is not None:
                column_width_dict[xl.utils.get_column_letter(head_column_th    )] = width   # 第n層  親側辺

            width = self._settings_obj.dictionary['column_width_of_child_side_edge']
            if width is not None:
                column_width_dict[xl.utils.get_column_letter(head_column_th + 1)] = width   #        子側辺

            width = self._settings_obj.dictionary['column_width_of_node']
            if width is not None:
                column_width_dict[xl.utils.get_column_letter(head_column_th + 2)] = width   #        節

            head_column_th += 3


        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width


        # 行の高さ設定
        # height の単位はポイント。初期値 8。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
        row_height_dict = {}

        height = self._settings_obj.dictionary['row_height_of_header']
        if height is not None:
            row_height_dict[1] = height

        height = self._settings_obj.dictionary['row_height_of_lower_side_padding']
        if height is not None:
            row_height_dict[2] = height

        for row_number, height in row_height_dict.items():
            ws.row_dimensions[row_number].height = height


        # 第１行
        # ------
        # ヘッダー行にする
        row_th = 1

        # TODO 可変長ノード数への対応
        # NOTE データテーブルではなく、ビュー用途なので、テーブルとしての機能性は無視しています
        # A の代わりに {xl.utils.get_column_letter(1)} とも書ける
        ws[f'A{row_th}'] = 'No'
        self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'A{row_th}'], index=0)
        self._settings_obj.set_font_of_header_to(cell=ws[f'A{row_th}'], index=0)

        # 根側パディング
        # --------------
        self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'B{row_th}'], index=1)

        # 根
        # --
        ws[f'C{row_th}'] = 'Root'
        self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'C{row_th}'], index=1)
        self._settings_obj.set_font_of_header_to(cell=ws[f'C{row_th}'], index=1)


        flip = 0
        head_column_th = 4

        for node_th in range(1, self._table.analyzer.end_th_of_node):
            # 背景色、文字色
            self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'{xl.utils.get_column_letter(head_column_th    )}{row_th}'], index=flip)
            self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'{xl.utils.get_column_letter(head_column_th + 1)}{row_th}'], index=flip)
            self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'{xl.utils.get_column_letter(head_column_th + 2)}{row_th}'], index=flip)
            self._settings_obj.set_font_of_header_to(cell=ws[f'{xl.utils.get_column_letter(head_column_th + 2)}{row_th}'], index=flip)

            # 列名
            ws[f'{xl.utils.get_column_letter(head_column_th + 2)}{row_th}'] = nth(node_th)

            flip = (flip + 1) % 2
            head_column_th += 3


        # 葉側パディング
        # --------------
        target_column_th = self._table.analyzer.end_th_of_node * StyleControl.ONE_NODE_COLUMNS + 1
        column_letter = xl.utils.get_column_letter(target_column_th)
        cell_address = f'{column_letter}{row_th}'
        # 背景色、文字色
        self._settings_obj.set_bgcolor_of_header_to(cell=ws[cell_address], index=(flip + 1) % 2)   # 葉ノードと同じ色にする
        self._settings_obj.set_font_of_header_to(cell=ws[cell_address], index=(flip + 1) % 2)

        width = self._settings_obj.dictionary['column_width_of_leaf_side_padding']
        if width is not None:
            ws.column_dimensions[column_letter].width = width


        # 余り列
        # ------
        # 最終層以降の列
        column_name_of_leaf_entry = self._table.analyzer.get_column_name_of_last_node()
        is_remaining = False
        target_column_th = self._table.analyzer.end_th_of_node * StyleControl.ONE_NODE_COLUMNS + 2   # 空列を１つ挟む
        for column_name in self._table.df.columns:

            # ツリー区は無視
            if column_name == column_name_of_leaf_entry:
                #print(f'ツリー区 {row_th=}  {column_name=}')
                is_remaining = True
                continue

            elif is_remaining:
                cell_address = f'{xl.utils.get_column_letter(target_column_th)}{row_th}'
                #print(f'{cell_address=}  {row_th=}  {column_name=}')

                # 列名
                ws[cell_address].value = column_name
                # 背景色、文字色
                self._settings_obj.set_bgcolor_of_header_to(cell=ws[cell_address], index=flip)   # 葉ノードと同じ色にする
                self._settings_obj.set_font_of_header_to(cell=ws[cell_address], index=flip)

                flip = (flip + 1) % 2
                target_column_th += 1


        # 第２行
        # ------
        # 空行にする
        row_th = 2
        self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'A{row_th}'], index=0)   # 葉ノードと同じ色にする

        # ツリー構造図の背景色
        for column_th in range(2, target_column_th):
            column_letter = xl.utils.get_column_letter(column_th)
            self._settings_obj.set_bgcolor_of_tree_to(cell=ws[f'{column_letter}{row_th}'])


    def _on_each_record(self, next_row_number, next_record):
        """先読みで最初の１回を空振りさせるので、２件目から本処理です"""

        # 事前送り出し
        self._forward_cursor(next_record=next_record)

        if self._curr_record.no is None:
            if self._debug_write:
                # 最初のレコードは先読みのため、空回しします
                print(f"[{datetime.datetime.now()}] Pencil {self._curr_record.no} record  first record read later")


        else:
            # 変数名短縮
            ws = self._ws


            # ３行目～６行目
            # --------------
            # データは３行目から、１かたまり３行を使って描画する
            HEADER_HEIGHT = 3
            RECORD_HEIGHT = 3
            curr_row_number = next_row_number - 1
            row1_th = curr_row_number * RECORD_HEIGHT + HEADER_HEIGHT
            row2_th = curr_row_number * RECORD_HEIGHT + HEADER_HEIGHT + 1
            row3_th = curr_row_number * RECORD_HEIGHT + HEADER_HEIGHT + 2
            three_row_numbers = [row1_th, row2_th, row3_th]

            # 行の高さ設定
            # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            height = self._settings_obj.dictionary['row_height_of_upper_side_of_node']
            if height is not None:
                ws.row_dimensions[row1_th].height = height
            
            height = self._settings_obj.dictionary['row_height_of_lower_side_of_node']
            if height is not None:
                ws.row_dimensions[row2_th].height = height
            
            height = self._settings_obj.dictionary['row_height_of_node_spacing']
            if height is not None:
                ws.row_dimensions[row3_th].height = height


            ws[f'A{row1_th}'].value = self._curr_record.no
            self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'A{row1_th}'], index=0)
            self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'A{row2_th}'], index=0)
            self._settings_obj.set_bgcolor_of_header_to(cell=ws[f'A{row3_th}'], index=0)

            # 根側のパディング
            # ----------------
            cells = [
                ws[f'B{row1_th}'],
                ws[f'B{row2_th}'],
                ws[f'B{row3_th}'],
            ]
            for cell in cells:
                self._settings_obj.set_bgcolor_of_tree_to(cell=cell)


            def draw_edge(depth_th, three_column_names, three_row_numbers):
                """
                Parameters
                ----------
                depth_th : int
                    第何層。根層は 0
                """

                cn1 = three_column_names[0]
                cn2 = three_column_names[1]
                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]


                # ツリー構造図の背景色
                # --------------------
                cells = [
                    ws[f'{cn1}{row1_th}'],
                    ws[f'{cn1}{row2_th}'],
                    ws[f'{cn1}{row3_th}'],
                    ws[f'{cn2}{row1_th}'],
                    ws[f'{cn2}{row2_th}'],
                    ws[f'{cn2}{row3_th}'],
                    ws[f'{cn3}{row1_th}'],
                    ws[f'{cn3}{row2_th}'],
                    ws[f'{cn3}{row3_th}'],
                ]
                for cell in cells:
                    self._settings_obj.set_bgcolor_of_tree_to(cell=cell)

                # ↑ノードが無くても背景色は塗る必要がある


                nd = self._curr_record.node_at(depth_th=depth_th)

                if nd is None or pd.isnull(nd.text):
                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Edge) {self._curr_record.no} record > {nth(depth_th)} layer  Empty cell")
                    return


                # 自件と前件を比較して、根から自ノードまで、ノードテキストが等しいか？
                if LineByLineLeafModel.is_same_path_as_avobe(
                        curr_record=self._curr_record,
                        prev_record=self._prev_record,
                        depth_th=depth_th):

                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Edge) {self._curr_record.no} record > {nth(depth_th)} layer  │")
                    
                    # 垂直線
                    #
                    #   |    leftside_border
                    # ..+..  
                    #   |    leftside_border
                    #   |    leftside_border
                    #
                    self._settings_obj.set_leftside_border_to_vertical(cell=ws[f'{cn2}{row1_th}'])
                    self._settings_obj.set_leftside_border_to_vertical(cell=ws[f'{cn2}{row2_th}'])
                    self._settings_obj.set_leftside_border_to_vertical(cell=ws[f'{cn2}{row3_th}'])
                    return


                # 子ノードへの接続は４種類の線がある
                #
                # (1) ─字
                #   .    under_border
                # ...__  
                #   .    None
                #   .    None
                #
                # (2) ┬字
                #   .    under_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (3) ├字
                #   |    l_letter_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (4) └字
                #   |    l_letter_border
                # ..+__  
                #   .    None
                #   .    None
                #
                kind = LineByLineLeafModel.get_kind_of_edge(
                        prev_record=self._prev_record,
                        curr_record=self._curr_record,
                        next_record=self._next_record,
                        depth_th=depth_th)

                if kind == '─字':
                    self._settings_obj.set_border_to_parent_horizontal(cell=ws[f'{cn1}{row1_th}'])
                    self._settings_obj.set_under_border_to_child_horizontal(cell=ws[f'{cn2}{row1_th}'])
                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Edge) {self._curr_record.no} record > {nth(depth_th)} layer  ─ {nd.edge_text}")
                
                elif kind == '┬字':
                    self._settings_obj.set_border_to_parent_downward(cell=ws[f'{cn1}{row1_th}'])
                    self._settings_obj.set_under_border_to_child_downward(cell=ws[f'{cn2}{row1_th}'])
                    self._settings_obj.set_leftside_border_to_child_downward(cell=ws[f'{cn2}{row2_th}'])
                    self._settings_obj.set_leftside_border_to_child_downward(cell=ws[f'{cn2}{row3_th}'])
                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Edge) {self._curr_record.no} record > {nth(depth_th)} layer  ┬ {nd.edge_text}")

                elif kind == '├字':
                    self._settings_obj.set_l_letter_border_to_child_rightward(cell=ws[f'{cn2}{row1_th}'])
                    self._settings_obj.set_leftside_border_to_child_rightward(cell=ws[f'{cn2}{row2_th}'])
                    self._settings_obj.set_leftside_border_to_child_rightward(cell=ws[f'{cn2}{row3_th}'])
                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Edge) {self._curr_record.no} record > {nth(depth_th)} layer  ├ {nd.edge_text}")

                elif kind == '└字':
                    self._settings_obj.set_l_letter_border_to_child_upward(cell=ws[f'{cn2}{row1_th}'])
                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Edge) {self._curr_record.no} record > {nth(depth_th)} layer  └ {nd.edge_text}")
                
                else:
                    raise ValueError(f"{kind=}")
                

                # ２列目：エッジ・テキスト
                ws[f'{cn2}{row1_th}'].value = nd.edge_text


            def draw_node(depth_th, three_column_names, three_row_numbers):
                """節を描きます

                Parameters
                ----------
                node : NodeInRecord
                    節
                depth_th : int
                    第何層。根層は 0
                """

                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]

                nd = self._curr_record.node_at(depth_th=depth_th)

                if nd is None or pd.isnull(nd.text) or LineByLineLeafModel.is_same_path_as_avobe(
                        curr_record=self._curr_record,
                        prev_record=self._prev_record,
                        depth_th=depth_th):

                    if self._debug_write:
                        print(f"[{datetime.datetime.now()}] Pencil(Node) {self._curr_record.no} record > {nth(depth_th)} layer  Empty cell")

                    # ツリー構造図の背景色
                    # --------------------
                    cells = [
                        ws[f'{cn3}{row1_th}'],
                        ws[f'{cn3}{row2_th}'],
                        ws[f'{cn3}{row3_th}'],
                    ]
                    for cell in cells:
                        self._settings_obj.set_bgcolor_of_tree_to(cell=cell)

                    return



                if self._debug_write:
                    print(f"[{datetime.datetime.now()}] Pencil(Node) {self._curr_record.no} record > {nth(depth_th)} layer  □ {nd.text}")
                
                ws[f'{cn3}{row1_th}'].value = nd.text
                self._settings_obj.set_alignment_of_node_to(cell=ws[f'{cn3}{row1_th}'])
                self._settings_obj.set_bgcolor_of_node_to(cell=ws[f'{cn3}{row1_th}'])
                self._settings_obj.set_border_of_upside_node(cell=ws[f'{cn3}{row1_th}'])

                self._settings_obj.set_bgcolor_of_node_to(cell=ws[f'{cn3}{row2_th}'])
                self._settings_obj.set_border_of_downside_node(cell=ws[f'{cn3}{row2_th}'])

                self._settings_obj.set_bgcolor_of_tree_to(cell=ws[f'{cn3}{row3_th}'])      # ツリー構造図の背景色

                # セル結合していいなら
                if not self._settings_obj.dictionary['do_not_merge_cells']:
                    #print(f"セル結合 {cn3}{row1_th}:{cn3}{row2_th}")
                    ws.merge_cells(f'{cn3}{row1_th}:{cn3}{row2_th}')

                    # FIXME 文字も数も区別なく左に揃えていいものか？
                    #self._settings_obj.set_alignment_left_center(cell=ws[f'{cn3}{row1_th}'])
                    self._settings_obj.set_alignment_center_center(cell=ws[f'{cn3}{row1_th}'])


            # 第０層
            # ------
            depth_th = 0
            column_letter = xl.utils.get_column_letter(3)   # 'C'
            if depth_th < self._table.analyzer.end_th_of_node:
                draw_node(depth_th=depth_th, three_column_names=[None, None, column_letter], three_row_numbers=three_row_numbers)


            # 第１～最終層
            # ------------
            for depth_th in range(1, self._table.analyzer.end_th_of_node):
                head_column_th = depth_th * StyleControl.ONE_NODE_COLUMNS + 1

                if depth_th < self._table.analyzer.end_th_of_node:
                    # 第1層は 'D', 'E', 'F'、以降、後ろにずれていく
                    column_letter_list = [
                        xl.utils.get_column_letter(head_column_th),
                        xl.utils.get_column_letter(head_column_th + 1),
                        xl.utils.get_column_letter(head_column_th + 2),
                    ]
                    draw_edge(depth_th=depth_th, three_column_names=column_letter_list, three_row_numbers=three_row_numbers)
                    draw_node(depth_th=depth_th, three_column_names=column_letter_list, three_row_numbers=three_row_numbers)


            column_name_of_last_node = self._table.analyzer.get_column_name_of_last_node()


            # 葉側のパディング
            # ----------------
            column_th = StyleControl.get_target_column_th(source_table=self._table, column_name=column_name_of_last_node) + 1
            column_letter = xl.utils.get_column_letter(column_th)
            cells = [
                ws[f'{column_letter}{row1_th}'],
                ws[f'{column_letter}{row2_th}'],
                ws[f'{column_letter}{row3_th}'],
            ]
            for cell in cells:
                self._settings_obj.set_bgcolor_of_tree_to(cell=cell)      # ツリー構造図の背景色



            # 余り列
            # ------
            # 最終層以降の列
            is_remaining = False
            for column_name in self._table.df.columns:
                if column_name == column_name_of_last_node:
                    is_remaining = True
                    continue

                elif is_remaining:
                    # TODO キャッシュを作れば高速化できそう
                    target_column_th = StyleControl.get_target_column_th(source_table=self._table, column_name=column_name)
                    column_letter = xl.utils.get_column_letter(target_column_th)

                    #print(f'{row1_th=}  {column_name=}')
                    ws[f'{column_letter}{row1_th}'].value = self._table.df.at[curr_row_number + 1, column_name]

                    # 罫線
                    self._settings_obj.set_upper_border_of_remaining_cell(cell=ws[f'{column_letter}{row1_th}'])
                    self._settings_obj.set_middle_border_of_remaining_cell(cell=ws[f'{column_letter}{row2_th}'])
                    self._settings_obj.set_lower_border_of_remaining_cell(cell=ws[f'{column_letter}{row3_th}'])

                    # ツリー構造図の背景色
                    cells = [
                        ws[f'{column_letter}{row1_th}'],
                        ws[f'{column_letter}{row2_th}'],
                        ws[f'{column_letter}{row3_th}'],
                    ]
                    for cell in cells:
                       self._settings_obj.set_bgcolor_of_tree_to(cell=cell)

                    # セル結合していいなら
                    if not self._settings_obj.dictionary['do_not_merge_cells']:
                        ws.merge_cells(f'{column_letter}{row1_th}:{column_letter}{row3_th}')

                        # FIXME 文字も数も区別なく左に揃えていいものか？
                        #self._settings_obj.set_alignment_left_center(cell=ws[f'{column_letter}{row1_th}'])
                        self._settings_obj.set_alignment_center_center(cell=ws[f'{column_letter}{row1_th}'])


class TreeEraser():
    """要らない罫線を消す"""


    def __init__(self, table, ws, settings_obj, debug_write=False):
        """初期化
        
        Parameters
        ----------
        table : Table
            ツリーテーブル
        ws : openpyxl.Worksheet
            ワークシート
        settings_obj : Settings
            各種設定
        debug_write : bool
            デバッグライト
            DEBUG_TIPS: デバッグライトをオンにして、コンソールにログを表示すると不具合を調査しやすくなります
        """
        self._table = table
        self._ws = ws
        self._settings_obj = settings_obj
        self._debug_write = debug_write


    def render(self):
        """描画"""

        # 指定の列の左側の垂直の罫線を見ていく
        column_th = 5
        for node_th in range(1, self._table.analyzer.end_th_of_node):
            self._erase_unnecessary_border_by_column(column_letter=xl.utils.get_column_letter(column_th))
            column_th += 3


    def _erase_unnecessary_border_by_column(self, column_letter):
        """不要な境界線を消す"""

        # 変数名の短縮
        ws = self._ws


        # 最後に見つけた、セルの左辺に罫線がなく、下辺に太い罫線がある行をリセット
        row_th_of_prev_last_underline = -1
        row_th_of_last_underline = -1


        # 第3行から
        row_th = 3
        while row_th <= ws.max_row: # 最終行まで全部見る

            # 前行のセルには、左辺と可変に太い罫線があったか？
            prerow_l_letter = False

            while True: # 仕切り直しの１セット
                shall_break = False

                currow_l_letter = False

                # 罫線を確認
                #
                #   .
                # ..+--  下向きの罫線が最後に出た箇所を調べる
                #   |
                #
                border = ws[f'{column_letter}{row_th}'].border
                if border is not None:
                    # セルの左辺に太い罫線が引かれており...
                    if border.left is not None and border.left.style == 'thick':
                        # セルの下辺にも太い罫線が引かれていれば、'└' 字か '├' 字のどちらかだ
                        if border.bottom is not None and border.bottom.style == 'thick':
                            row_th_of_prev_last_underline = row_th_of_last_underline
                            row_th_of_last_underline = row_th
                            currow_l_letter = True
                            if self._debug_write:
                                # 左側と下側に罫線。 '└' 字か '├' 字のどちらかだ。アンダーラインが第何行か覚えておく
                                print(f"[{datetime.datetime.now()}] Eraser {column_letter}{row_th} Border on the left and bottom. '└' or '├'. Memory underline ({row_th_of_last_underline} row) (Preview {row_th_of_prev_last_underline} row)")

                        # 左辺に罫線。次行へ読み進めていく
                        else:
                            if self._debug_write:
                                print(f"[{datetime.datetime.now()}] Eraser {column_letter}{row_th} Border on the left")

                    # セルの左辺に太い罫線が引かれていない
                    else:
                        # "└"字。［ラスト・シブリング］なので、最後に見つけた左辺に罫線のないアンダーラインのことは忘れて仕切り直し
                        if prerow_l_letter:
                            row_th_of_prev_last_underline = -1
                            row_th_of_last_underline = -1
                            shall_break = True
                            if self._debug_write:
                                print(f"[{datetime.datetime.now()}] Eraser {column_letter}{row_th} Last sibling. Forget underline row")

                        # セルの下辺に太い罫線が引かれていたら、つながっていない垂線だ。それが第何行か覚えておいて仕切り直す
                        elif border.bottom is not None and border.bottom.style == 'thick':
                            row_th_of_prev_last_underline = row_th_of_last_underline
                            row_th_of_last_underline = row_th
                            shall_break = True
                            if self._debug_write:
                                print(f"[{datetime.datetime.now()}] Eraser {column_letter}{row_th} Memory underline ({row_th_of_last_underline} row) (Preview {row_th_of_prev_last_underline} row)")

                        # セルの左辺にも、下辺にも、太い罫線が引かれていなければ、罫線は尻切れトンボになっている。仕切り直し
                        else:
                            row_th_of_prev_last_underline = row_th_of_last_underline
                            row_th_of_last_underline = row_th
                            shall_break = True
                            if self._debug_write:
                                print(f"[{datetime.datetime.now()}] Eraser {column_letter}{row_th} No border on the left and bottom. no connect line. Memory ({row_th_of_last_underline} row) (Preview {row_th_of_prev_last_underline} row)")


                row_th += 1

                prerow_l_letter = currow_l_letter

                if shall_break:
                    break


            # 消しゴムを掛ける
            start_row_to_erase = row_th_of_prev_last_underline + 1
            end_row_to_erase = row_th_of_last_underline

            if row_th_of_last_underline != -1 and 0 < start_row_to_erase and start_row_to_erase < end_row_to_erase:

                if self._debug_write:
                    print(f"[{datetime.datetime.now()}] Eraser {column_letter}_ Erase {start_row_to_erase} to {end_row_to_erase - 1} row...")

                for row_th_to_erase in range(start_row_to_erase, end_row_to_erase):
                    # 消すか、見え消しにするか切り替えられるようにしておく
                    self._settings_obj.set_striked_border(cell=ws[f'{column_letter}{row_th_to_erase}'])

        if self._debug_write:
            print(f"[{datetime.datetime.now()}] Eraser {column_letter}{row_th} finished (EOL {ws.max_row})")
