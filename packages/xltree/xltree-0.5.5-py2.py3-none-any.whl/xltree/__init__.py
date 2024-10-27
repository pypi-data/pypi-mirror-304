import gc
import os
import datetime
import openpyxl as xl
from .models.tree_structure import Forest
from .settings import Settings
from .worksheet_handle import WorksheetHandle


def planting():
    """森を作成します"""
    return Forest()


def prepare_workbook(target, mode, settings={}, debug_write=False):
    """ワークブックを用意します
    
    Returns
    -------
    target : str
        ワークブック（.xlsx）へのファイルパス
    mode : str
        既存のワークブックが有ったときの挙動。 'w' は新規作成して置換え、 'a' は追記
    settings : dict
        各種設定    
    debug_write : bool
        デバッグライト
    """

    wb = None

    # 既存のファイルが有ったときの挙動
    if os.path.isfile(target):
        # 既存のファイルへ追記
        if mode == 'a':
            if debug_write:
                print(f"[{datetime.datetime.now()}] `{target}` file exists, read.")

            # ワークブックを開く
            wb = xl.load_workbook(filename=target)
        
        elif mode == 'w':
            pass
        
        else:
            raise ValueError(f"unsupported {mode=}")

                
    # ワークブックを新規生成
    if wb is None:
        if debug_write:
            print(f"[{datetime.datetime.now()}] `{target}` file not exists, create.")

        wb = xl.Workbook()


    return WorkbookHandle(target=target, mode=mode, wb=wb, settings=settings, debug_write=debug_write)


class XltreeInSrc():
    """テストで使う仕組み

    テストでは

        import xltree as tr

    のように書けないので、
    テストでは以下のように書く

        from src.xltree import xltree_in_src as tr
    
    """


    @staticmethod
    def planting():
        """グローバル関数の planting() を呼び出す
        """
        global planting

        return planting()


    @staticmethod
    def prepare_workbook(target, mode, settings={}, debug_write=False):
        """グローバル関数の prepare_workbook() を呼び出す
        
        Returns
        -------
        target : str
            ワークブック（.xlsx）へのファイルパス
        mode : str
            既存のワークブックが有ったときの挙動。 'w' は新規作成して置換え、 'a' は追記
        settings : dict
            各種設定    
        debug_write : bool
            デバッグライト
        """
        global prepare_workbook

        return prepare_workbook(target=target, mode=mode, settings=settings, debug_write=debug_write)


xltree_in_src = XltreeInSrc()


class WorkbookHandle():
    """ワークブックへの対応"""


    def __init__(self, target, mode, wb, settings={}, debug_write=False):
        """初期化

        Parameters
        ----------
        target : str
            ワークブック（.xlsx）へのファイルパス
        mode : str
            既存のワークブックが有ったときの挙動。 'w' は新規作成して置換え、 'a' は追記
        settings : dict
            各種設定
        """
        self._wb_file_path = target
        self._mode = mode
        self._wb = wb
        self._settings_obj = Settings(dictionary=settings)
        self._debug_write = debug_write
        self._ws = None


    def __enter__(self):
        return self


    def __exit__(self, exc_type, exc_value, traceback):
        del self._wb_file_path
        del self._mode
        del self._wb
        del self._settings_obj
        del self._debug_write
        del self._ws

        # メモリ解放
        gc.collect()


    @property
    def workbook_file_path(self):
        return self._wb_file_path


    def prepare_worksheet(self, target, based_on, debug_write=False):
        """ワークシートの用意"""

        # ワークシートの準備
        if not self.exists_sheet(target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] create `{target}` sheet...")

            self._wb.create_sheet(target)


        ws = self._wb[target]


        # ワークシート制御の生成
        return WorksheetHandle.instantiate(target=target, based_on=based_on, ws=ws, settings_obj=self._settings_obj, debug_write=debug_write)



    def render_worksheet(self, target, based_on, debug_write=False):
        """ワークシートへ木構造図を描画

        Parameters
        ----------
        target : str
            シート名
        based_on : str
            CSVファイルパス
        debug_write : bool
            デバッグライト
        """

        # ワークシート制御の生成
        wsh = WorksheetHandle.instantiate(target=target, based_on=based_on, ws=self._ws, settings_obj=self._settings_obj, debug_write=debug_write)

        # 木の描画
        wsh.render_tree()


    def remove_worksheet(self, target):
        """存在すれば、指定のワークシートの削除。存在しなければ無視

        Parameters
        ----------
        target : str
            シート名
        """

        if self.exists_sheet(target=target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] remove `{target}` sheet...")

            self._wb.remove(self._wb[target])


    def save_workbook(self):
        """ワークブックの保存"""

        if self._debug_write:
            print(f"[{datetime.datetime.now()}] save `{self._wb_file_path}` file...")

        # ワークブックの保存            
        self._wb.save(self._wb_file_path)


    def exists_sheet(self, target):
        """シートの存在確認
        
        Parameters
        ----------
        target : str
            シート名
        """
        return target in self._wb.sheetnames


    def get_worksheet(self, sheet_name):
        """ワークシートの取得"""
        return self._wb[sheet_name]
