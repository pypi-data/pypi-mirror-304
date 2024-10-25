from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment


class Settings():
    """各種設定"""


    def __init__(self, dictionary=None):
        """初期化
        
        Parameters
        ----------
        dictionary : dict
            設定

            列の幅設定。width はだいたい 'ＭＳ Ｐゴシック' サイズ11 の半角英文字の個数
            * `column_width_of_no` - A列の幅。no列
            * `column_width_of_root_side_padding` - B列の幅。ツリー構造図の根側パディング
            * `column_width_of_leaf_side_padding` - ツリー構造図の葉側パディング
            * `column_width_of_node` - 例：C, F, I ...列の幅。ノードの箱の幅
            * `column_width_of_parent_side_edge` - 例：D, G, J ...列の幅。エッジの水平線のうち、親ノードの方
            * `column_width_of_child_side_edge` - 例：E, H, K ...列の幅。エッジの水平線のうち、子ノードの方

            行の高さ設定。height の単位はポイント。既定値 8。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            * `row_height_of_header` - 第１行。ヘッダー
            * `row_height_of_lower_side_padding` - 第２行。ツリー構造図の軸の番号が小さい側パティング
            * `row_height_of_upper_side_of_node` - ノードの上側のセルの高さ
            * `row_height_of_lower_side_of_node` - ノードの下側のセルの高さ
            * `row_height_of_node_spacing` - ノード間の高さ

            * 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)

            背景色関連
            * `bgcolor_of_tree` - ツリー構造図の背景
            * `bgcolor_of_header_1` - ヘッダーの背景色その１
            * `bgcolor_of_header_2` - ヘッダーの背景色その２
            * `bgcolor_of_node` - 背景色

            文字色関連
            * `fgcolor_of_header_1` - ヘッダーの文字色その１
            * `fgcolor_of_header_2` - ヘッダーの文字色その２

            文字寄せ関連
            * `horizontal_alignment_of_node` - 文字の水平方向の寄せ。規定値 None。'left', 'fill', 'centerContinuous', 'center', 'right', 'general', 'justify', 'distributed' のいずれか。指定しないなら None
            * `vertical_alignment_of_node` - 文字の垂直方向の寄せ。規定値 None。'bottom', 'center', 'top', 'justify', 'distributed' のいずれか。指定しないなら None

            その他の操作
            * `do_not_merge_cells` - セル結合しないなら真
        """

        # 既定のディクショナリー
        # いわゆる settings
        self._dictionary = {
            # 列の幅
            #
            #   ［列幅の自動調整］機能を付けたので、文字が入る箇所は規定値はナンにします。
            #   キーは存在させたいので、コメントアウトしないでください
            #
            'column_width_of_no':                    None,
            'column_width_of_root_side_padding':        3,
            'column_width_of_leaf_side_padding':        3,
            'column_width_of_node':                  None,
            'column_width_of_parent_side_edge':         2,
            'column_width_of_child_side_edge':       None,

            # 行の高さ
            'row_height_of_header':                    13,
            'row_height_of_lower_side_padding':        13,
            'row_height_of_upper_side_of_node':        13,
            'row_height_of_lower_side_of_node':        13,
            'row_height_of_node_spacing':               6,

            # 背景色関連
            'bgcolor_of_tree':                   'FFFFFF',
            'bgcolor_of_header_1':               'CCCCCC',
            'bgcolor_of_header_2':               '333333',
            'bgcolor_of_node':                   'FFFFCC',

            # 文字色関連
            'fgcolor_of_header_1':               '111111',
            'fgcolor_of_header_2':               'EEEEEE',

            # 文字寄せ関連
            'horizontal_alignment_of_node':          None,
            'vertical_alignment_of_node':            None,

            # その他の操作
            'do_not_merge_cells':                   False,      # セル結合しないなら真
        }

        # 上書き
        if dictionary is not None:
            for key, value in dictionary.items():
                self._dictionary[key] = value


        # フォント関連
        # ------------
        self._font_of_header_list = []

        color = self.dictionary['fgcolor_of_header_1']
        if color is not None:
            self._font_of_header_list.append(Font(color=color))
        
        color = self.dictionary['fgcolor_of_header_2']
        if color is not None:
            self._font_of_header_list.append(Font(color=color))


        # 罫線関連
        # --------

        # エッジの罫線
        #
        #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
        #   色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
        #
        BLACK = '000000'
        side = Side(style='thick', color=BLACK)

        # DEBUG_TIPS: 罫線に色を付けると、デバッグしやすいです
        if True:
            red_side = Side(style='thick', color=BLACK)
            orange_side = Side(style='thick', color=BLACK)
            green_side = Side(style='thick', color=BLACK)
            blue_side = Side(style='thick', color=BLACK)
            cyan_side = Side(style='thick', color=BLACK)
        else:
            red_side = Side(style='thick', color='FF0000')
            orange_side = Side(style='thick', color='FFCC00')
            green_side = Side(style='thick', color='00FF00')
            blue_side = Side(style='thick', color='0000FF')
            cyan_side = Side(style='thick', color='00FFFF')

        # ─字  赤
        self._border_to_parent_horizontal = Border(bottom=red_side)
        self._under_border_to_child_horizontal = Border(bottom=red_side)
        # │字  緑
        self._leftside_border_to_vertical = Border(left=green_side)
        # ┬字  青
        self._border_to_parent_downward = Border(bottom=blue_side)
        self._under_border_to_child_downward = Border(bottom=blue_side)
        self._leftside_border_to_child_downward = Border(left=blue_side)
        # ├字  青緑
        self._l_letter_border_to_child_rightward = Border(left=cyan_side, bottom=cyan_side)
        self._leftside_border_to_child_rightward = Border(left=cyan_side)
        # └字  橙
        self._l_letter_border_to_child_upward = Border(left=orange_side, bottom=orange_side)

        # DEBUG_TIPS: デバッグ時は、罫線を消すのではなく、灰色に変えると見やすいです
        if True:
            # 罫線無し
            self._striked_border = None
        else:
            # 罫線
            #
            #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
            #   色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
            #
            # 見え消し用（デバッグに使う）
            striked_side = Side(style='thick', color='DDDDDD')
            # 見え消し用の罫線
            self._striked_border = Border(left=striked_side)


        # ノードの罫線
        #
        #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
        #
        side = Side(style='thick', color='000000')
        self._border_of_upside_node = Border(top=side, left=side, right=side)
        self._border_of_downside_node = Border(bottom=side, left=side, right=side)


        # 余り列
        side = Side(style='thin', color='111111')
        self._upper_border_of_remaining_cell = Border(top=side, left=side, right=side)
        self._middle_border_of_remaining_cell = Border(left=side, right=side)
        self._lower_border_of_remaining_cell = Border(bottom=side, left=side, right=side)


        # 背景色関連
        # ----------
        self._list_of_bgcolor_of_header = []

        color = self.dictionary['bgcolor_of_header_1']
        if color is not None:
            self._list_of_bgcolor_of_header.append(PatternFill(patternType='solid', fgColor=color))
        else:
            self._list_of_bgcolor_of_header.append(None)
        
        color = self.dictionary['bgcolor_of_header_2']
        if color is not None:
            self._list_of_bgcolor_of_header.append(PatternFill(patternType='solid', fgColor=color))
        else:
            self._list_of_bgcolor_of_header.append(None)

        # ツリー構造図の背景色
        color = self.dictionary['bgcolor_of_tree']
        if color is not None:
            self._bgcolor_of_tree = PatternFill(patternType='solid', fgColor=color)

        # ノードの背景色
        color = self.dictionary['bgcolor_of_node']
        if color is not None:
            self._bgcolor_of_node = PatternFill(patternType='solid', fgColor=color)


        # 文字寄せ関連
        # ------------
        #
        #   horizontal は 'distributed', 'fill', 'general', 'center', 'centerContinuous', 'justify', 'right', 'left' のいずれかから選ぶ
        #   vertical は 'center', 'top', 'bottom', 'justify', 'distributed' のいずれかから選ぶ
        #   TODO Excelでは文字は left、数は right など規定値が型によって違うが、そこまで細かく設定を分けてない。分けるか？
        #
        self._alignment_left_center = Alignment(horizontal='left', vertical='center')
        self._alignment_center_center = Alignment(horizontal='center', vertical='center')

        horizontal = self.dictionary['horizontal_alignment_of_node']
        vertical = self.dictionary['vertical_alignment_of_node']
        if horizontal is not None and vertical is not None:
            self._alignment_of_node = Alignment(horizontal=horizontal, vertical=vertical)
        elif horizontal is not None:
            self._alignment_of_node = Alignment(horizontal=horizontal)
        elif vertical is not None:
            self._alignment_of_node = Alignment(vertical=vertical)
        else:
            self._alignment_of_node = None


    @property
    def dictionary(self):
        return self._dictionary


    def set_alignment_left_center(self, cell):
        cell.alignment = self._alignment_left_center


    def set_alignment_center_center(self, cell):
        cell.alignment = self._alignment_center_center


    def set_striked_border(self, cell):
        # FIXME None にするという動作。どう対称性を取る？
        #if self._striked_border is not None:
        cell.border = self._striked_border


    def set_border_of_upside_node(self, cell):
        if self._border_of_upside_node is not None:
            cell.border = self._border_of_upside_node


    def set_border_of_downside_node(self, cell):
        if self._border_of_downside_node is not None:
            cell.border = self._border_of_downside_node


    def set_border_to_parent_horizontal(self, cell):
        if self._border_to_parent_horizontal is not None:
            cell.border = self._border_to_parent_horizontal


    def set_under_border_to_child_horizontal(self, cell):
        if self._under_border_to_child_horizontal is not None:
            cell.border = self._under_border_to_child_horizontal


    def set_leftside_border_to_vertical(self, cell):
        if self._leftside_border_to_vertical is not None:
            cell.border = self._leftside_border_to_vertical


    def set_border_to_parent_downward(self, cell):
        if self._border_to_parent_downward is not None:
            cell.border = self._border_to_parent_downward


    def set_under_border_to_child_downward(self, cell):
        if self._under_border_to_child_downward is not None:
            cell.border = self._under_border_to_child_downward


    def set_leftside_border_to_child_downward(self, cell):
        if self._leftside_border_to_child_downward is not None:
            cell.border = self._leftside_border_to_child_downward


    def set_l_letter_border_to_child_rightward(self, cell):
        if self._l_letter_border_to_child_rightward is not None:
            cell.border = self._l_letter_border_to_child_rightward


    def set_leftside_border_to_child_rightward(self, cell):
        if self._leftside_border_to_child_rightward is not None:
            cell.border = self._leftside_border_to_child_rightward


    def set_l_letter_border_to_child_upward(self, cell):
        if self._l_letter_border_to_child_upward is not None:
            cell.border = self._l_letter_border_to_child_upward


    def set_upper_border_of_remaining_cell(self, cell):
        if self._upper_border_of_remaining_cell is not None:
            cell.border = self._upper_border_of_remaining_cell


    def set_middle_border_of_remaining_cell(self, cell):
        if self._middle_border_of_remaining_cell is not None:
            cell.border = self._middle_border_of_remaining_cell


    def set_lower_border_of_remaining_cell(self, cell):
        if self._lower_border_of_remaining_cell is not None:
            cell.border = self._lower_border_of_remaining_cell


    def set_bgcolor_of_header_to(self, cell, index):
        if self._list_of_bgcolor_of_header[index] is not None:
            cell.fill = self._list_of_bgcolor_of_header[index]


    def set_font_of_header_to(self, cell, index):
        if self._font_of_header_list[index] is not None:
            cell.font = self._font_of_header_list[index]


    def set_bgcolor_of_tree_to(self, cell):
        if self._bgcolor_of_tree is not None:
            cell.fill = self._bgcolor_of_tree


    def set_bgcolor_of_node_to(self, cell):
        if self._bgcolor_of_node is not None:
            cell.fill = self._bgcolor_of_node


    def set_alignment_of_node_to(self, cell):
        if self._alignment_of_node is not None:
            cell.alignment = self._alignment_of_node


class SettingsOfNode():
    """TODO ノード個別の設定"""


    def __init__(self, dictionary=None):
        """初期化
        
        Parameters
        ----------
        dictionary : dict
            設定

            * 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)

            色関連
            * `bgcolor` - 背景色
            * `fgcolor` - 文字色

            文字寄せ関連
            * `horizontal_alignment` - 文字の水平方向の寄せ。規定値 None。'left', 'fill', 'centerContinuous', 'center', 'right', 'general', 'justify', 'distributed' のいずれか。指定しないなら None
            * `vertical_alignment_` - 文字の垂直方向の寄せ。規定値 None。'bottom', 'center', 'top', 'justify', 'distributed' のいずれか。指定しないなら None
        """

        # 既定のディクショナリー
        self._dictionary = {

            # 色関連
            'bgcolor':                     'FFFFCC',
            'fgcolor':                     None,

            # 文字寄せ関連
            'horizontal_alignment':        None,
            'vertical_alignment':          None,
        }

        # 上書き
        if dictionary is not None:
            for key, value in dictionary.items():
                self._dictionary[key] = value


    @property
    def dictionary(self):
        return self._dictionary
